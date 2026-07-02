# -*- coding: utf-8 -*-
"""
hopyo_parser.py — 단가산출근거 호표별 일작업량 파서

배경
----
설계자가 단가산출근거에 직접 산정한 시공량(시간당 작업량 Q)을 호표번호로
1:1 매칭해서, 가이드라인보다 우선 적용하기 위한 데이터 소스를 만든다.

버전 이력
---------
v1: 형식1 "Q = ... = X 단위/HR" (대문자만) → 106개 호표 커버.
v2: 아래 2개 케이스 추가 → 278개 호표로 확대.
    (a) 대소문자 버그 수정: 시트 안에 "/HR", "/hr", "/Hr" 등 표기가
        섞여 있었는데 v1 정규식이 대문자 "/HR"만 매칭해서 소문자
        표기(217개 호표)를 전부 놓치고 있었음 → re.IGNORECASE로 수정.
    (b) 역수 형태(시간/단위) 신규 지원: 일부 호표는
        "Q = X hr/단위" (예: 0.16 hr/본, "본 하나 만드는 데 걸리는 시간")
        형태로 되어 있어 기존 "단위/hr" 패턴과 정반대. 이 경우
        일작업량 = 8 / X 로 계산 (기존 X*8과 다른 공식).
        58개 호표가 이 형태.

처리 대상 형식
--------------
1) "Q = ... = X.XX 단위/hr" (대소문자 무관)  → 일작업량 = X * 8
   "Q1 = ... = X.XX 단위/hr" (항발 등, 변수명 Q/Q1 무관)
2) "Q = ... = X.XX hr/단위" (역수형: 단위당 소요시간) → 일작업량 = 8 / X
같은 블록에 두 형식이 섞여 있진 않다고 가정(각 호표는 1가지 형식만 사용).
한 블록에 여러 줄이 있으면 처음 발견되는 것을 채택.

전제 조건
---------
- 워크북은 반드시 `data_only=True` 로 로드되어야 함.
  단가산출근거 시트의 Q값 텍스트는 Excel 수식 결과로 만들어지므로,
  data_only=False 로 열면 수식 원문이 들어와 정규식이 매칭되지 않는다.
- 호표 시작 라인 '제 N호표'와 Q 산식 라인 둘 다 열 인덱스 1에 있음.

반환 형식
---------
{호표번호(int): (일작업량(float, 본/일 등), 단위(str))}

호표 228 단일 예외
------------------
'열연강판 항타 및 항발' 호표는 한 블록에 Q=(항타)와 Q1=(항발) 두 줄이 있다.
현재 구현은 문서 순서상 첫 줄(항타)을 사용한다 → 보수적(작은 값) 선택.

표준 사용
---------
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    hopyo_daily = parse_hopyo_daily_amounts(wb)
    # 예) hopyo_daily[74] == (34.48, '본')

CLI 검증
--------
    python hopyo_parser.py <엑셀경로>
"""

from __future__ import annotations

import re
import sys
from typing import Dict, Tuple

# 열 인덱스 1의 '제 N호표' 패턴 (블록 시작 표시)
_HOPYO_BLOCK_RE = re.compile(r"제\s*(\d+)\s*호표")

# 형식1(정방향): 'Q = ... = X 단위/hr' (대소문자 무관: /HR, /hr, /Hr 전부 매칭)
_Q_FORWARD_RE = re.compile(r"=\s*([\d.]+)\s*([가-힣㎡㎥]+)/hr", re.IGNORECASE)

# 형식2(역수): 'Q = ... = X hr/단위' (단위 하나 만드는 데 걸리는 시간)
_Q_INVERSE_RE = re.compile(r"=\s*([\d.]+)\s*hr/([가-힣㎡㎥]+)", re.IGNORECASE)

# 시간당 → 일당 환산 계수 (8시간 작업)
_HR_TO_DAY = 8


def parse_hopyo_daily_amounts(wb) -> Dict[int, Tuple[float, str]]:
    """
    단가산출근거 시트를 1회 스캔해서 호표번호 → (일작업량, 단위) dict 반환.

    Parameters
    ----------
    wb : openpyxl.Workbook
        반드시 ``data_only=True`` 로 로드된 워크북.

    Returns
    -------
    dict[int, tuple[float, str]]
        {호표번호: (일작업량, 단위)}. 두 형식 다 못 찾은 호표는 포함되지 않음.
    """
    if wb is None or "단가산출근거" not in wb.sheetnames:
        return {}

    ws = wb["단가산출근거"]
    rows = [r for r in ws.iter_rows(values_only=True)]

    # 1) 각 호표 시작 행 인덱싱 (첫 출현만 채택)
    starts: Dict[int, int] = {}
    for i, r in enumerate(rows):
        c1 = r[1] if len(r) > 1 else None
        if isinstance(c1, str):
            m = _HOPYO_BLOCK_RE.search(c1)
            if m:
                n = int(m.group(1))
                if n not in starts:
                    starts[n] = i
    if not starts:
        return {}

    # 2) 각 호표 블록(다음 호표 시작 직전까지) 안에서 첫 Q 라인 추출
    #    정방향(단위/hr)을 우선 탐색하고, 없으면 역수(hr/단위)를 시도.
    ordered = sorted(starts.items(), key=lambda kv: kv[1])
    result: Dict[int, Tuple[float, str]] = {}
    for idx, (n, s) in enumerate(ordered):
        end = ordered[idx + 1][1] if idx + 1 < len(ordered) else len(rows)

        found_fwd = None
        found_inv = None
        for j in range(s, end):
            c1 = rows[j][1] if len(rows[j]) > 1 else None
            if not isinstance(c1, str):
                continue
            if found_fwd is None:
                m = _Q_FORWARD_RE.search(c1)
                if m:
                    found_fwd = (float(m.group(1)), m.group(2))
            if found_fwd is None and found_inv is None:
                mi = _Q_INVERSE_RE.search(c1)
                if mi:
                    found_inv = (float(mi.group(1)), mi.group(2))
            if found_fwd is not None:
                break  # 정방향이 우선이므로 찾으면 바로 종료

        if found_fwd is not None:
            val, unit = found_fwd
            result[n] = (round(val * _HR_TO_DAY, 4), unit)
        elif found_inv is not None:
            val, unit = found_inv
            if val > 0:
                result[n] = (round(_HR_TO_DAY / val, 4), unit)

    return result


def parse_hopyo_daily_amounts_from_path(path: str) -> Dict[int, Tuple[float, str]]:
    """파일 경로에서 직접 로드해서 파싱. 통합 외 단독 실행/테스트용 헬퍼."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return parse_hopyo_daily_amounts(wb)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# CLI 검증: python hopyo_parser.py <엑셀경로>
# ---------------------------------------------------------------------------

# (호표번호, 기대 일작업량, 허용오차, 설명)
_VERIFY_POINTS = [
    # v1 기존 검증 (회귀 확인 — 이 값들은 절대 안 바뀌어야 함)
    (1,   32.56, 0.05, "H-PILE 천공후항타 H=3.55m"),
    (40,  58.48, 0.05, "SHEET PILE 천공후항타 H=3.8m  (검증치 58.5)"),
    (46,  54.00, 0.05, "SHEET PILE 천공후항타 H=4.48m (검증치 54.0)"),
    (74,  34.48, 0.05, "SHEET PILE 천공후항타 H=9.5m  (검증치 34.5)"),
    (15,  60.00, 0.05, "H-PILE 항발 H=3.55m"),
    (25,  51.92, 0.05, "H-PILE 항발 H=6.21m"),
    (110, 39.60, 0.05, "SHEET PILE 항발 H=9.5m"),
    (228, 30.32, 0.05, "열연강판 항타 및 항발 (첫 Q라인=항타)"),
    # v2 신규 검증 (대소문자 버그 수정으로 커버된 항목)
    (111, 75.46 * 8, 0.5, "가물막이 설치 및 해체 (소문자 ㎥/hr)"),
    (117, 18.53 * 8, 0.5, "되메우기(B=1.5m미만,모래) (소문자 ㎥/hr)"),
    # v2 신규 검증 (역수 형태 hr/단위)
    (129, 50.00, 0.05, "말뚝박기용 천공(열연강판) H=2.9m (역수: 0.16hr/본)"),
    (130, 50.00, 0.05, "말뚝박기용 천공(오거비트) H=2.94m (역수: 0.16hr/본)"),
]


def _main(argv):
    import warnings
    warnings.filterwarnings("ignore")  # openpyxl Print area 경고 억제

    if len(argv) < 2:
        print("사용법: python hopyo_parser.py <엑셀경로>", file=sys.stderr)
        return 2
    path = argv[1]

    print(f"[로드] {path}")
    result = parse_hopyo_daily_amounts_from_path(path)
    print(f"[파싱] 형식1+역수형 통합 보유 호표: {len(result)}개")

    # 단위 분포
    from collections import Counter
    units = Counter(u for _, u in result.values())
    print(f"[단위] {dict(units)}")

    # 검증 포인트 평가
    print()
    print("=== 검증 포인트 ===")
    ok = 0
    fail = 0
    for n, expected, tol, desc in _VERIFY_POINTS:
        if n not in result:
            print(f"  ✗ 호표 {n:>3}  {desc}  → 매칭 안 됨 (기대 {expected})")
            fail += 1
            continue
        daily, unit = result[n]
        diff = abs(daily - expected)
        status = "✓" if diff <= tol else "✗"
        print(f"  {status} 호표 {n:>3}  {desc}")
        print(f"        실제 {daily}{unit}, 기대 {expected}, 오차 {diff:.4f} (허용 {tol})")
        if diff <= tol:
            ok += 1
        else:
            fail += 1

    print()
    print(f"=== 결과: {ok}/{len(_VERIFY_POINTS)} 통과, {fail} 실패 ===")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    sys.exit(_main(sys.argv))