import streamlit as st
import pandas as pd
import openpyxl
from hopyo_parser import parse_hopyo_daily_amounts
import math
import re
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO

# 가이드라인 데이터 및 기상 데이터 import
try:
    from guideline_data import GUIDELINE_APPENDIX_FULL, GUIDELINE_APPENDIX
    from weather_data import (
        RAIN_DAYS, COLD_DAYS, HOT_DAYS, REGION_MAPPING,
        get_total_non_work_days, get_monthly_breakdown
    )
    from holiday_data import (
        LEGAL_HOLIDAYS, get_legal_holidays, get_total_holidays,
        calc_overlap_days, get_total_non_work_days_with_holidays,
        get_holiday_breakdown_monthly
    )
    REGIONS = list(REGION_MAPPING.keys())
except ImportError as e:
    print(f"Import 실패: {e}")
    # 파일이 없을 경우 기본값 사용
    GUIDELINE_APPENDIX_FULL = {}
    GUIDELINE_APPENDIX = {}
    RAIN_DAYS = {}
    COLD_DAYS = {}
    HOT_DAYS = {}
    REGION_MAPPING = {"서울": "서울"}
    REGIONS = ["서울"]
    LEGAL_HOLIDAYS = {}
    def get_total_non_work_days(region, start_date, end_date, check_rain=True, check_cold=True, check_hot=True):
        return 0
    def get_monthly_breakdown(region, start_date, end_date, check_rain=True, check_cold=True, check_hot=True):
        return []
    def get_legal_holidays(year, month):
        return 0
    def get_total_holidays(start_date, end_date):
        return 0
    def calc_overlap_days(a, b, c):
        return 0
    def get_total_non_work_days_with_holidays(*args, **kwargs):
        return {"total": 0, "weather": 0, "holidays": 0, "overlap": 0, "formula": ""}
    def get_holiday_breakdown_monthly(start_date, end_date):
        return []

st.set_page_config(page_title="상하수도 공기산정", layout="wide", initial_sidebar_state="expanded")

# ══════════════════════════════════════════════════════════════
# 공종 키워드 매핑
# ══════════════════════════════════════════════════════════════
KEYWORD_MAP_DETAIL = {
    "포장복구": ["포장복구","아스팔트포장","아스팔트+콘크리트포장","콘크리트포장","보도포장","인도포장","보조기층","택코팅","프라임코팅","기층","표층","차선도색","노면절삭","아스팔트 노면 절삭","절삭후","보도블럭","과속방지턱","줄눈설치","줄눈"],
    "굴착공": ["터파기","굴착","줄파기","착공","시굴","포장깨기","포장절단"],
    "관부설공": ["관 부설","관부설","PE다중벽관","고강성PVC","주철관","GRP관",
                 "유리섬유복합관","흄관","이중벽관","강관부설","콘크리트관"],
    "되메우기": ["되메우기","뒤채움","복토","성토"],
    "맨홀공": ["맨홀","우수받이","집수정","토실","슬라이딩"],
    "배수설비": ["배수설비","빗물받이"],
    "추진공": ["추진공","추진관","추진"],
}

SKIP_NAMES = [
    "남천지구","동부지구","신설오수관로","간선관로","지선관로",
    "순공사비","배수설비공사","토공","관로공","구조물공","포장공",
    "추진공","부대공","안전관리비","환경보전비","소계","합계","계",
]

PIPE_EXCLUDE = ["절단","이형관","하차비","단관","마감캡","추진관"]
MACHINE_BASED = ["터파기","굴착","되메우기","모래기초","모래부설","모래,관기초"]

# ══════════════════════════════════════════════════════════════
# 가이드라인 부록 데이터 (대폭 확장)
# ══════════════════════════════════════════════════════════════
GUIDELINE_APPENDIX = {
    # 포장공
    "아스팔트포장 절단": {"daily": 1000, "unit": "m"},
    "아스팔트포장깨기 (B.H0.4㎥)": {"daily": 515, "unit": "㎡"},
    "아스팔트포장깨기 (B.H0.7㎥)": {"daily": 1047, "unit": "㎡"},
    "콘크리트포장 절단": {"daily": 500, "unit": "m"},
    "콘크리트포장 깨기": {"daily": 300, "unit": "㎡"},
    "아스팔트포장(기층)": {"daily": 800, "unit": "㎡"},
    "아스팔트포장(택코팅)": {"daily": 2000, "unit": "㎡"},
    "아스팔트포장(프라임코팅)": {"daily": 2000, "unit": "㎡"},
    "아스팔트포장(표층)": {"daily": 1000, "unit": "㎡"},
    "택코팅": {"daily": 2000, "unit": "㎡"},
    "프라임코팅": {"daily": 2000, "unit": "㎡"},
    "기층": {"daily": 800, "unit": "㎡"},
    "표층": {"daily": 1000, "unit": "㎡"},
    
    # 터파기
    "터파기(토사:육상) B/H 0.4㎥": {"daily": 260, "unit": "㎥"},
    "터파기(토사:육상) B/H 0.7㎥": {"daily": 530, "unit": "㎥"},
    "터파기(암:육상) B/H 0.4㎥": {"daily": 130, "unit": "㎥"},
    "터파기(암:육상) B/H 0.7㎥": {"daily": 265, "unit": "㎥"},
    
    # 되메우기 (폭별 추가)
    "되메우기(진동롤러) 2.5ton": {"daily": 600, "unit": "㎥"},
    "되메우기(진동롤러) 4.0ton": {"daily": 950, "unit": "㎥"},
    "되메우기(진동콤팩터)": {"daily": 400, "unit": "㎥"},
    "되메우기(B=1.5~2.5m)": {"daily": 450, "unit": "㎥"},
    "되메우기(B=2.5~4.0m)": {"daily": 650, "unit": "㎥"},
    "되메우기(B=4.0m이상)": {"daily": 850, "unit": "㎥"},
    
    # 관부설
    "관부설(D200)": {"daily": 5, "unit": "본/일"},
    "관부설(D300)": {"daily": 4, "unit": "본/일"},
    "관부설(D450)": {"daily": 3, "unit": "본/일"},
    "관부설(D600)": {"daily": 2.5, "unit": "본/일"},
    "관부설(D800)": {"daily": 2, "unit": "본/일"},
    "관부설(D1000)": {"daily": 1.5, "unit": "본/일"},
    "관부설(D1200)": {"daily": 1.2, "unit": "본/일"},
    
    # 맨홀공 (대폭 확장)
    "원형맨홀 Φ1200": {"daily": 2.5, "unit": "개소/일"},
    "원형맨홀 Φ1500": {"daily": 2.5, "unit": "개소/일"},
    "각형맨홀 1800×2400": {"daily": 1.5, "unit": "개소/일"},
    "우수받이": {"daily": 5, "unit": "개소/일"},
    "조립식 PC맨홀": {"daily": 3, "unit": "개소/일"},
    "GRP맨홀": {"daily": 2, "unit": "개소/일"},
    "조립식맨홀설치(소형)": {"daily": 8, "unit": "개소/일"},
    "조립식맨홀 상부구체": {"daily": 10, "unit": "개소/일"},
    "조립식맨홀 연직구체": {"daily": 12, "unit": "개소/일"},
    "조립식맨홀 하부구체": {"daily": 8, "unit": "개소/일"},
    "맨홀뚜껑설치": {"daily": 20, "unit": "개소/일"},
    "맨홀뚜껑": {"daily": 20, "unit": "개소/일"},
    
    # 배수설비
    "빗물받이": {"daily": 5, "unit": "개소/일"},
    "집수받이": {"daily": 4, "unit": "개소/일"},
    "우수토실": {"daily": 3, "unit": "개소/일"},
    "배수설비": {"daily": 4, "unit": "개소/일"},
    
    # 추진공
    "추진설비공": {"daily": 1, "unit": "개소/일"},
    "사토": {"daily": 100, "unit": "㎥/일"},
    "추진마감벽설치": {"daily": 3, "unit": "개소/일"},
    "추진마감벽": {"daily": 3, "unit": "개소/일"},
    "천공홀 되메우기": {"daily": 500, "unit": "m"},
    
    # 가시설공
    "조립식 간이 흙막이": {"daily": 50, "unit": "㎡/일"},
    "H-PILE 항타": {"daily": 8, "unit": "본/일"},
    
    # 기타
    "보조기층": {"daily": 500, "unit": "㎡/일"},
    "모래기초": {"daily": 400, "unit": "㎥/일"},
}

# ══════════════════════════════════════════════════════════════
# 표준품셈 노무량
# ══════════════════════════════════════════════════════════════
def get_excavation_labor(spec_str):
    labor_table = {
        0.4: {"인/m3": 0.130},
        0.7: {"인/m3": 0.085},
        1.0: {"인/m3": 0.070},
    }
    if "0.4" in spec_str or "B.H0.4" in spec_str or "B/H 0.4" in spec_str:
        return labor_table[0.4]
    elif "0.7" in spec_str or "B.H0.7" in spec_str or "B/H 0.7" in spec_str:
        return labor_table[0.7]
    elif "1.0" in spec_str or "B.H1.0" in spec_str or "B/H 1.0" in spec_str:
        return labor_table[1.0]
    return {"인/m3": 0.085}

def get_pipe_labor(diameter):
    pipe_labor = {
        200: {"합계": 0.396},
        300: {"합계": 0.494},
        450: {"합계": 0.653},
        600: {"합계": 0.792},
        800: {"합계": 0.990},
        1000: {"합계": 1.188},
        1200: {"합계": 1.386},
    }
    closest = min(pipe_labor.keys(), key=lambda x: abs(x - diameter))
    return pipe_labor.get(closest, {"합계": 0.5})

def is_machine_based(name):
    return any(kw in name for kw in MACHINE_BASED)

def extract_diameter(spec_str):
    patterns = [r'D\s*[=＝]?\s*(\d+)',r'Φ\s*(\d+)',r'φ\s*(\d+)',
                r'(\d{2,4})\s*(?:mm|㎜)',r'[D=]?(\d{2,4})']
    for pat in patterns:
        m = re.search(pat, spec_str)
        if m:
            val = int(m.group(1))
            if 50 <= val <= 3000:
                return val
    return None

def calc_days_priority(name, spec, qty, crews=3, item_unit=""):
    """
    우선순위:
    0. 수동입력
    1. 단가산출근거 호표 Q값 (호표번호 1:1 매칭)
    2. 가이드라인 부록
    3. 표준품셈 Man-day
    4. 단가산출근거 Q값 (항목명 기반)
    
    item_unit: 실제 항목의 단위 (M, 개소, 본 등)
    """
    if not qty or qty <= 0:
        return 0, "-", "-"

    # 0순위: 수동입력 (최우선)
    try:
        if "manual_rates" in st.session_state:
            manual_key = f"{name}|{spec}"  # 이름+규격으로 키 생성
            if manual_key in st.session_state["manual_rates"]:
                manual_data = st.session_state["manual_rates"][manual_key]
                daily_val = manual_data.get("daily", 0)
                unit = manual_data.get("unit", "")
                if daily_val > 0:
                    days = math.ceil(qty / (daily_val * crews))
                    return days, f"{daily_val:.1f}{unit}", "수동입력"
    except Exception:
        pass

    # 1순위: 단가산출근거 호표 Q값 (호표번호로 1:1 매칭; 형식1=Q=/HR 기반)
    try:
        hopyo_map = st.session_state.get("hopyo_by_item", {})
        hopyo_daily = st.session_state.get("hopyo_daily", {})
        hopyo_no = hopyo_map.get((name, spec))
        if hopyo_no is not None and hopyo_no in hopyo_daily:
            daily_val, unit = hopyo_daily[hopyo_no]
            if daily_val and daily_val > 0:
                days = math.ceil(qty / (daily_val * crews))
                return days, f"{daily_val:.1f}{unit}", "단가산출근거(호표)"
    except Exception:
        pass

    # 1순위: 가이드라인
    try:
        # 정확한 매칭 시도
        full_name = f"{name} {spec}".strip()
        
        # GUIDELINE_APPENDIX_FULL 우선 사용 (확장판)
        guideline_data = GUIDELINE_APPENDIX_FULL if GUIDELINE_APPENDIX_FULL else GUIDELINE_APPENDIX
        
        # 띄어쓰기 제거 및 괄호 제거한 버전 준비
        full_name_no_space = full_name.replace(" ", "").replace("(", "").replace(")", "")
        name_no_space = name.replace(" ", "").replace("(", "").replace(")", "")
        
        for key, val in guideline_data.items():
            matched = False
            key_no_space = key.replace(" ", "").replace("(", "").replace(")", "")
            
            # 매칭 조건 (우선순위)
            # 1. 정확한 전체 매칭 (띄어쓰기/괄호 무시)
            if key_no_space == full_name_no_space or key_no_space == name_no_space:
                matched = True
            
            # 2. 가이드라인 키가 항목명에 포함 (띄어쓰기/괄호 무시)
            elif key_no_space in full_name_no_space or key_no_space in name_no_space:
                matched = True
            
            # 3. 항목명이 가이드라인 키에 포함 (띄어쓰기/괄호 무시)
            elif name_no_space in key_no_space:
                matched = True
            
            # 4. 원본 문자열 매칭 (띄어쓰기 있는 버전)
            elif key == full_name or key == name or key in full_name or key in name:
                matched = True
            
            # 5. 핵심 키워드 매칭 (특수 케이스)
            # "조립식맨홀설치" → "조립식맨홀" 매칭
            elif "조립식맨홀" in key and "조립식맨홀" in name:
                matched = True
            elif "맨홀뚜껑" in key and "맨홀뚜껑" in name:
                matched = True
            # 🔥 수정: 추진공은 "강관" + "추진" 조합만 매칭 (오매칭 방지)
            # 원래 코드: elif "추진" in key and "추진" in name and len(key) > 2:
            # 문제: "추진" 키워드만으로 다른 항목과 오매칭되어 1850일 폭발
            elif "추진" in key and "추진" in name and "강관" in key and "강관" in (name + spec):
                matched = True
            
            if matched:
                base_daily = val.get("daily", 0)
                unit = val.get("unit", "")
                
                # ⚠️ 의심스러운 가이드라인 스킵 (이상 값 방지)
                # 1. unit="일" 같이 명확하지 않은 단위
                # 2. daily가 1 미만으로 너무 작은 값 (1m/일 미만은 비정상)
                unit_clean = unit.strip().lower()
                if unit_clean in ["일", "day", "days", ""]:
                    # 단위가 "일"이면 의미 불명확 → 매칭 안 됨으로 처리
                    continue
                if base_daily > 0 and base_daily < 1:
                    # 1일에 1단위 미만은 너무 작음 → 매칭 안 됨으로 처리
                    continue
                
                # ⚠️ 단위 불일치 체크
                if item_unit and unit and base_daily > 0:
                    # 단위 정규화
                    item_unit_clean = item_unit.strip().lower().replace(" ", "")
                    guideline_unit_clean = unit.split("/")[0].strip().lower().replace(" ", "")
                    
                    # M/일 vs 개소, 본/일 vs M 같은 불일치 감지
                    unit_mismatch = False
                    if "m" in guideline_unit_clean or "ｍ" in guideline_unit_clean:
                        if item_unit_clean not in ["m", "ｍ", "m3", "㎥"]:
                            unit_mismatch = True
                    elif "본" in guideline_unit_clean:
                        if item_unit_clean not in ["본", "ea", "개"]:
                            unit_mismatch = True
                    elif "개소" in guideline_unit_clean or "ea" in guideline_unit_clean:
                        if item_unit_clean not in ["개소", "ea", "개", "본"]:
                            unit_mismatch = True
                    
                    if unit_mismatch:
                        # 단위 불일치 → 이 가이드라인은 스킵
                        continue
                
                if base_daily > 0:
                    if is_machine_based(name):
                        days = math.ceil(qty / (base_daily * crews))
                        label = f"{base_daily}{unit}"  # 조수 제거
                    else:
                        days = math.ceil(qty / (base_daily * crews))
                        label = f"{base_daily}{unit}"  # 조수 제거
                    return days, label, "가이드라인"
        
        # 관부설 직경별 매칭
        if any(kw in name for kw in ["관 부설","관부설","고강성PVC","PE다중벽","이중벽관","주철관","GRP관"]):
            dia = extract_diameter(spec)
            if dia:
                pipe_rates = {200:5, 300:4, 450:3, 600:2.5, 800:2, 1000:1.5, 1200:1.2}
                closest = min(pipe_rates.keys(), key=lambda x: abs(x - dia))
                daily = pipe_rates[closest]
                days = math.ceil(qty / (daily * crews))
                return days, f"{daily}본/일", "가이드라인"  # 조수 제거
    except Exception:
        pass

    # 2순위: 표준품셈
    try:
        manday = 0
        if any(kw in name for kw in ["터파기","굴착","줄파기"]) and "운반" not in name:
            info = get_excavation_labor(spec)
            rate = info.get("인/m3")
            if rate:
                manday = rate * qty

        pipe_kws = ["관 부설","관부설","이중벽관","주철관","흄관","콘크리트관",
                    "GRP관","유리섬유복합관","파형강관","PE다중벽","고강성PVC","강관부설"]
        if any(kw in name for kw in pipe_kws) and not manday:
            dia = extract_diameter(spec)
            if dia:
                info = get_pipe_labor(dia)
                rate = info.get("합계")
                if rate:
                    manday = rate * qty

        if manday > 0:
            days = math.ceil(manday / (8 * crews))
            return days, f"{round(manday/qty,3)}인/단위", "표준품셈"  # 조수 제거
    except Exception:
        pass
    
    # 3순위: 단가산출근거
    try:
        if "dangagun_cache" in st.session_state:
            cache = st.session_state["dangagun_cache"]
            
            # 항목명 + 규격으로 매칭 시도
            full_name = f"{name} {spec}".strip()
            
            for cached_name, info in cache.items():
                # 정확한 매칭 우선
                if cached_name == full_name or cached_name in full_name or full_name in cached_name:
                    # hourly 값 (시간당)
                    if "hourly" in info:
                        hourly_val = info.get("hourly", 0)
                        unit = info.get("unit", "")
                        if hourly_val > 0:
                            daily_val = hourly_val * 8
                            days = math.ceil(qty / (daily_val * crews))
                            return days, f"{daily_val:.1f}{unit.replace('/Hr','/일')}", "단가산출근거"  # 조수 제거
                    
                    # daily 값 (1일 작업량)
                    elif "daily" in info:
                        daily_val = info.get("daily", 0)
                        unit = info.get("unit", "")
                        if daily_val > 0:
                            days = math.ceil(qty / (daily_val * crews))
                            return days, f"{daily_val:.1f}{unit}", "단가산출근거"  # 조수 제거
                
                # 항목명만으로도 매칭 시도
                if name in cached_name or cached_name in name:
                    if "hourly" in info:
                        hourly_val = info.get("hourly", 0)
                        unit = info.get("unit", "")
                        if hourly_val > 0:
                            daily_val = hourly_val * 8
                            days = math.ceil(qty / (daily_val * crews))
                            return days, f"{daily_val:.1f}{unit.replace('/Hr','/일')}", "단가산출근거"  # 조수 제거
                    
                    elif "daily" in info:
                        daily_val = info.get("daily", 0)
                        unit = info.get("unit", "")
                        if daily_val > 0:
                            days = math.ceil(qty / (daily_val * crews))
                            return days, f"{daily_val:.1f}{unit}", "단가산출근거"  # 조수 제거
    except Exception:
        pass

    # 4순위: 매칭 안 됨 → 수동입력 필요
    return 0, "⚠️ 수동입력", "매칭 안 됨"

# ══════════════════════════════════════════════════════════════
# 비작업일수
# ══════════════════════════════════════════════════════════════
HOLIDAYS_DB = {
    2025:{1:8,2:4,3:7,4:4,5:6,6:6,7:4,8:6,9:4,10:9,11:5,12:5},
    2026:{1:5,2:7,3:6,4:4,5:7,6:5,7:4,8:7,9:7,10:7,11:5,12:5},
}
RAIN = {1:0,2:1,3:2,4:3,5:4,6:6,7:8,8:7,9:5,10:3,11:2,12:1}

def get_kr_holidays(year):
    m = HOLIDAYS_DB.get(year, {})
    holidays = set()
    for month, count in m.items():
        for day in range(1, count + 1):
            try:
                holidays.add(datetime(year, month, day).date())
            except:
                pass
    return holidays

def calc_completion_date(start, work_days):
    current, worked = start, 0
    kr_holidays = get_kr_holidays(start.year) | get_kr_holidays(start.year + 1)
    while worked < work_days:
        if current.weekday() == 6 or current in kr_holidays or current.day % 30 < RAIN[current.month]:
            current += timedelta(days=1)
            continue
        worked += 1
        current += timedelta(days=1)
    return current - timedelta(days=1)

# ══════════════════════════════════════════════════════════════
# 엑셀 파서
# ══════════════════════════════════════════════════════════════
def parse_by_keyword(file):
    # 🔥 디버그: 파일 로그
    import datetime
    log_file = "debug_log.txt"
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"\n{'='*60}\n")
        f.write(f"파싱 시작: {datetime.datetime.now()}\n")
        f.write(f"{'='*60}\n")
    
    # 🔥 디버그: 파싱 시작
    print(f"\n{'🔥'*30}")
    print(f"📂 parse_by_keyword 시작")
    print(f"{'🔥'*30}\n")
    
    wb = openpyxl.load_workbook(file, data_only=True)  # read_only=False로 변경
    skip_sheets = ["목차","안내","INITIAL","초기","index"]
    priority = ["설계내역서","내역서","공사비내역서"]
    target_sheet = None
    
    for p in priority:
        if p in wb.sheetnames:
            target_sheet = p
            break
    if not target_sheet:
        for sname in wb.sheetnames:
            if any(sk in sname for sk in skip_sheets):
                continue
            if "내역" in sname:
                target_sheet = sname
                break
    if not target_sheet:
        for sname in wb.sheetnames:
            if not any(sk in sname for sk in skip_sheets):
                target_sheet = sname
                break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]

    ws = wb[target_sheet]
    col_info = {"시트명": target_sheet}
    header_row = None
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        row_str = " ".join([str(c) for c in row if c])
        if any(k in row_str for k in ["공종","품명","세부품명","명칭","내역"]):
            header_row = row_idx
            break
    
    if not header_row:
        header_row = 1
    
    col_info["헤더행"] = header_row
    
    # ══════════════════════════════════════════════════════════════
    # 지구 경계 찾기 (로마숫자)
    # ══════════════════════════════════════════════════════════════
    districts = {}
    roman_nums = ['Ⅰ', 'Ⅱ', 'Ⅲ', 'Ⅳ', 'Ⅴ', 'Ⅵ', 'Ⅶ', 'Ⅷ', 'Ⅸ', 'Ⅹ']
    
    all_rows_raw = list(ws.iter_rows(min_row=1, values_only=False))
    
    for row_idx, row in enumerate(all_rows_raw):
        a_val = str(row[0].value or "").strip()
        b_val = str(row[1].value or "").strip() if len(row) > 1 else ""
        
        if a_val in roman_nums:
            districts[a_val] = {
                'name': b_val,
                'start_row': row_idx,
                'end_row': None
            }
    
    # 지구별 end_row 설정
    district_keys = sorted(districts.keys(), key=lambda x: roman_nums.index(x))
    for i, key in enumerate(district_keys):
        if i + 1 < len(district_keys):
            next_key = district_keys[i + 1]
            districts[key]['end_row'] = districts[next_key]['start_row'] - 1
        else:
            districts[key]['end_row'] = len(all_rows_raw) - 1
    
    col_info["districts"] = districts
    
    # ══════════════════════════════════════════════════════════════
    # 지구별 데이터 파싱
    # ══════════════════════════════════════════════════════════════
    results = []
    
    for district, info in districts.items():
        district_rows = all_rows_raw[info['start_row']:info['end_row']+1]
        
        for local_idx, row in enumerate(district_rows):
            row_idx = info['start_row'] + local_idx
            
            # values_only=False이므로 .value 접근
            gong_jong_val = row[0].value
            name_val = row[1].value if len(row) > 1 else None
            spec_val = row[2].value if len(row) > 2 else None
            qty_val = row[3].value if len(row) > 3 else None
            unit_val = row[4].value if len(row) > 4 else None
            
            gong_jong = str(gong_jong_val).strip() if gong_jong_val else ""
            name = str(name_val).strip() if name_val else ""
            spec = str(spec_val).strip() if spec_val else ""
            unit = str(unit_val).strip() if unit_val else ""
            
            if not name or any(skip in name for skip in SKIP_NAMES):
                continue
            
            try:
                qty = float(qty_val) if qty_val else 0
            except:
                qty = 0
            
            if qty <= 0:
                continue
            
            group = "기타"
            for grp, keywords in KEYWORD_MAP_DETAIL.items():
                if any(kw in name for kw in keywords):
                    group = grp
                    break
            
            if group == "관부설공" and any(ex in name for ex in PIPE_EXCLUDE):
                group = "기타"
            
            detail_spec = spec
            if not detail_spec and name:
                spec_match = re.search(r'\([^)]+\)', name)
                if spec_match:
                    detail_spec = spec_match.group(0)
            
            # 호표 참조 추출 (행 전체 스캔; 통상 비고열에 '산근 N호표' 텍스트가 있음)
            hopyo_num = None
            for cell in row:
                v = cell.value
                if isinstance(v, str):
                    m_ref = re.search(r'산근\s*(\d+)\s*호표', v)
                    if m_ref:
                        hopyo_num = int(m_ref.group(1))
                        break

            results.append({
                "row_idx": row_idx,
                "district": district,
                "district_name": info['name'],
                "gong_jong": gong_jong,
                "group": group,
                "name": name,
                "spec": detail_spec,
                "qty": qty,
                "unit": unit,
                "amount": row[5].value if len(row) > 5 else 0,
                "labor": row[6].value if len(row) > 6 else 0,
                "hopyo": hopyo_num,
            })
    
    # 원본 순서로 정렬 (row_idx 기준)
    results.sort(key=lambda x: x["row_idx"])
    
    # 중복 제거 (전체 데이터 기준)
    merged = {}
    for r in results:
        key = (r["name"], r["spec"])
        if key not in merged:
            merged[key] = dict(r)
        else:
            merged[key]["qty"] = (merged[key].get("qty") or 0) + (r.get("qty") or 0)
            merged[key]["amount"] = (merged[key].get("amount") or 0) + (r.get("amount") or 0)
            merged[key]["labor"] = (merged[key].get("labor") or 0) + (r.get("labor") or 0)
            # 호표는 같은 (name, spec)이면 동일하다고 가정. 만일 첫 행에 비어있고 뒤에 채워졌다면 채워준다.
            if merged[key].get("hopyo") is None and r.get("hopyo") is not None:
                merged[key]["hopyo"] = r["hopyo"]
    
    return list(merged.values()), col_info

# ══════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════
st.sidebar.header("⚙️ 기본 설정")
st.sidebar.markdown("""
<div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
            padding: 16px 20px; border-radius: 10px; margin-bottom: 8px;'>
    <h3 style='color: white; margin: 0; font-size: 18px;'>🚧 공사 유형</h3>
</div>
""", unsafe_allow_html=True)
construction_type = st.sidebar.selectbox(
    "공사 유형 선택",
    ["하수관로", "상하수도 구조물", "관로+구조물"],
    index=0,
    label_visibility="collapsed",
    help="현재는 하수관로 위주로 동작합니다. 상하수도 구조물(하수처리시설·정수장·배수지·펌프장 등)은 추후 지원 예정.",
)
st.session_state["construction_type"] = construction_type

# 📋 워크플로우 가이드
st.sidebar.markdown("""
<div style='background: linear-gradient(135deg, #1e3a8a 0%, #1e40af 100%); 
            padding: 16px; border-radius: 10px; margin-bottom: 20px;'>
    <h3 style='color: white; margin: 0 0 12px 0; font-size: 16px;'>📋 작업 순서</h3>
    <ol style='color: #dbeafe; margin: 0; padding-left: 20px; font-size: 13px; line-height: 1.8;'>
        <li>📂 <strong>엑셀 인식</strong><br><span style='font-size: 11px; color: #93c5fd;'>설계내역서 업로드</span></li>
        <li>📝 <strong>수동입력</strong><br><span style='font-size: 11px; color: #93c5fd;'>매칭 안 된 항목 입력</span></li>
        <li>🌧 <strong>비작업일수</strong><br><span style='font-size: 11px; color: #93c5fd;'>기상·휴일 반영</span></li>
        <li>📋 <strong>공기산정</strong><br><span style='font-size: 11px; color: #93c5fd;'>투입조수·작업일수</span></li>
        <li>🔍 <strong>CP 분석</strong><br><span style='font-size: 11px; color: #93c5fd;'>주요공종 식별</span></li>
        <li>📄 <strong>보고서</strong><br><span style='font-size: 11px; color: #93c5fd;'>최종 결과 출력</span></li>
    </ol>
</div>
""", unsafe_allow_html=True)

st.sidebar.info("📅 **공사 시작일**은 TAB '비작업일수 계산기'에서 설정")
st.title("상하수도 공사기간 산정 시스템")
st.markdown("---")

tab2, tab6, tab4, tab1, tab3, tab5 = st.tabs([
    "📂 엑셀 내역서 인식",
    "📝 수동입력 관리",
    "🌧 비작업일수 계산기",
    "📋 공기산정",
    "🔍 주요공종 CP 분석",
    "📄 공기산정 보고서"
])

# ══════════════════════════════════════════════════════════════
# TAB 2
# ══════════════════════════════════════════════════════════════
with tab2:
    st.subheader("📂 엑셀 내역서 자동 인식")
    st.caption("도급 설계내역서 업로드 → 계층 구조 자동 파싱")

    uploaded = st.file_uploader("설계내역서 엑셀 (.xlsx)", type=["xlsx","xls"])

    if uploaded:
        # 프로그레스 바
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        try:
            status_text.text("📂 엑셀 파일 로드 중...")
            progress_bar.progress(20)
            
            with st.spinner("파싱 중..."):
                all_rows, col_info = parse_by_keyword(uploaded)
            
            progress_bar.progress(40)
            status_text.text("✅ 파싱 완료!")
            
            matched = [r for r in all_rows if r["group"] != "기타"]
            
            progress_bar.progress(60)
            status_text.text("🔍 계층 구조 분석 중...")
            
            if matched:
                status_text.text("📊 UI 생성 중...")
                progress_bar.progress(80)
                
                st.markdown("---")
                
                wb = openpyxl.load_workbook(uploaded, data_only=True)
                ws = wb['설계내역서'] if '설계내역서' in wb.sheetnames else wb.active
                
                # 단가산출근거 캐싱 (개선: 다양한 패턴 인식)
                dangagun_cache = {}
                if '단가산출근거' in wb.sheetnames:
                    ws_danga = wb['단가산출근거']
                    current_item = None
                    
                    for row in ws_danga.iter_rows(min_row=1, values_only=True):
                        row_text = " ".join([str(c) for c in row if c])
                        
                        # 항목명 추출 (규격 포함)
                        if row[1] and "/" in str(row[1]):
                            item_text = str(row[1]).strip()
                            if "/" in item_text:
                                current_item = item_text.split("/")[0].strip()
                        
                        # Q 값 추출 (다양한 패턴)
                        if current_item and "Q =" in row_text:
                            # 패턴 1: Q = 숫자 단위/HR
                            match1 = re.search(r'Q\s*=\s*([\d.]+)\s*([^\s]+/HR)', row_text, re.IGNORECASE)
                            if match1:
                                hourly_val = float(match1.group(1))
                                unit = match1.group(2).replace("HR", "Hr").replace("hr", "Hr")
                                dangagun_cache[current_item] = {"hourly": hourly_val, "unit": unit}
                                continue
                            
                            # 패턴 2: Q = 숫자/일 /8 Hr = 숫자 단위/Hr
                            match2 = re.search(r'=\s*([\d.]+)\s*([^\s/]+)/Hr', row_text, re.IGNORECASE)
                            if match2:
                                hourly_val = float(match2.group(1))
                                unit = match2.group(2) + "/Hr"
                                dangagun_cache[current_item] = {"hourly": hourly_val, "unit": unit}
                                continue
                        
                        # 1세트 = N일 패턴
                        if current_item and "세트" in row_text and "일" in row_text:
                            match3 = re.search(r'(\d+)\s*세트\s*=\s*([\d.]+)\s*일', row_text)
                            if match3:
                                sets = float(match3.group(1))
                                days = float(match3.group(2))
                                # 1일 = sets/days 세트
                                daily_val = sets / days
                                dangagun_cache[current_item] = {"daily": daily_val, "unit": "세트/일"}
                                continue
                
                st.session_state["dangagun_cache"] = dangagun_cache
                
                if dangagun_cache:
                    st.info(f"✅ 단가산출근거에서 {len(dangagun_cache)}개 항목 Q값 추출")
                
                # 단가산출근거 호표별 일작업량 캐시 (호표번호 1:1 매칭용; 형식1=Q=/HR 기반)
                st.session_state["hopyo_daily"] = parse_hopyo_daily_amounts(wb)
                if st.session_state["hopyo_daily"]:
                    st.info(f"✅ 단가산출근거 호표별 일작업량 {len(st.session_state['hopyo_daily'])}개 추출")
                
                # 계층 구조 파싱
                hierarchy = []
                st.session_state["hopyo_by_item"] = {}  # (name, spec) → 호표번호 (calc 1순위용, 매 파싱마다 초기화)
                current_category = None
                current_sub_category = None
                current_sub_sub_category = None  # 3단계 계층
                current_district = None  # 지구 추적
                
                roman_nums = ['Ⅰ', 'Ⅱ', 'Ⅲ', 'Ⅳ', 'Ⅴ', 'Ⅵ', 'Ⅶ', 'Ⅷ', 'Ⅸ', 'Ⅹ']
                
                for row in ws.iter_rows(min_row=1, values_only=True):
                    gong_jong = str(row[0]).strip() if row[0] else ""
                    name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    spec = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    
                    # 지구 변경 감지
                    if gong_jong in roman_nums:
                        # 이전 sub_category를 category에 추가
                        if current_sub_category and current_category:
                            is_already_in_list = any(s is current_sub_category for s in current_category['sub_categories'])
                            if not is_already_in_list:
                                current_category['sub_categories'].append(current_sub_category)
                        
                        # 지구 전환
                        current_district = gong_jong
                        current_sub_category = None  # sub 초기화
                        current_sub_sub_category = None  # sub_sub도 초기화
                        continue
                    
                    if re.match(r'^\d+\.\d+\.\d+$', gong_jong):
                        # 같은 level+name이면 기존 카테고리 재사용 (지구별 합산)
                        existing = next((c for c in hierarchy if c['level'] == gong_jong and c['name'] == name), None)
                        
                        if existing:
                            # 기존 카테고리 재사용
                            if current_category and current_category != existing:
                                if current_sub_category:
                                    current_category['sub_categories'].append(current_sub_category)
                                    current_sub_category = None
                            current_category = existing
                            current_sub_category = None
                            continue
                        else:
                            # 새 카테고리 생성
                            if current_category:
                                if current_sub_category:
                                    current_category['sub_categories'].append(current_sub_category)
                                    current_sub_category = None
                                if current_category.get('items') or current_category.get('sub_categories'):
                                    hierarchy.append(current_category)
                            
                            current_category = {
                                'level': gong_jong,
                                'name': name,
                                'items': [],
                                'sub_categories': []
                            }
                            current_sub_category = None
                            continue
                    
                    # (N) #숫자 형태는 구분자로 sub_sub_category 생성
                    # 예: (1) #1 추진, (2) #2 추진 등 - 각각 다른 구간이므로 독립적으로 계산
                    is_hash_separator = False
                    if re.match(r'^\(\d+\)$', gong_jong) and name and re.match(r'^#\d+', name):
                        is_hash_separator = True
                    elif re.match(r'^#\d+', gong_jong):
                        is_hash_separator = True
                    
                    if is_hash_separator:
                        print(f"🟣 (#N) 구분자 감지: gong_jong=[{gong_jong}] name=[{name}] → sub_sub_category로 생성")
                        # #1, #2... 를 sub_sub_category로 생성 (지구 정보 포함)
                        if current_sub_category:
                            current_sub_sub_category = {
                                'level': gong_jong,
                                'name': name,
                                'district': current_district,
                                'items': []
                            }
                            # sub_categories 리스트에 추가 (키 이름 통일!)
                            if 'sub_categories' not in current_sub_category:
                                current_sub_category['sub_categories'] = []
                            current_sub_category['sub_categories'].append(current_sub_sub_category)
                            print(f"  ✅ sub_sub 생성: parent=[{current_sub_category['name']}] level=[{gong_jong}] name=[{name}] district=[{current_district}]")
                        continue
                    # 1), 2) 패턴은 sub_category (2단계 계층)
                    elif re.match(r'^\d+\)$', gong_jong):
                        print(f"🔵 1) 패턴 감지: gong_jong=[{gong_jong}] name=[{name}]")
                        # 1), 2) 형태는 sub_category
                        if current_category:
                            # 이전 sub_sub 마무리
                            if current_sub_sub_category and current_sub_category:
                                is_already_in_list = any(s is current_sub_sub_category for s in current_sub_category['sub_categories'])
                                if not is_already_in_list:
                                    if 'sub_categories' not in current_sub_category:
                                        current_sub_category['sub_categories'] = []
                                    current_sub_category['sub_categories'].append(current_sub_sub_category)
                            current_sub_sub_category = None
                            
                            # 같은 level+name+district의 sub_category가 있으면 재사용
                            existing_sub = next((s for s in current_category['sub_categories'] 
                                               if s['level'] == gong_jong and s['name'] == name and s.get('district') == current_district), None)
                            
                            if existing_sub:
                                print(f"🔵 sub 재사용: category=[{current_category['name']}] level=[{gong_jong}] name=[{name}] district=[{current_district}]")
                                # 이전 sub가 있으면 append (기존과 다른 경우에만)
                                if current_sub_category and current_sub_category != existing_sub:
                                    is_already_in_list = any(s is current_sub_category for s in current_category['sub_categories'])
                                    if not is_already_in_list:
                                        print(f"  ⚠️ 이전 sub [{current_sub_category.get('name')}] append")
                                        current_category['sub_categories'].append(current_sub_category)
                                # 기존 sub로 교체
                                current_sub_category = existing_sub
                            else:
                                print(f"🟢 sub 생성: category=[{current_category['name']}] level=[{gong_jong}] name=[{name}] district=[{current_district}]")
                                # 현재 sub가 있으면 append
                                if current_sub_category:
                                    # 같은 객체가 이미 리스트에 있는지 체크 (객체 ID)
                                    is_already_in_list = any(s is current_sub_category for s in current_category['sub_categories'])
                                    if not is_already_in_list:
                                        current_category['sub_categories'].append(current_sub_category)
                                current_sub_category = {
                                    'level': gong_jong,
                                    'name': name,
                                    'items': [],
                                    'sub_categories': [],  # 3단계를 위한 sub_categories
                                    'district': current_district  # 지구 정보 추가
                                }
                        # continue를 제거하여 다음 행 계속 읽기
                    
                    # 항목 매칭
                    if current_category and not gong_jong and name:
                        # 엑셀 row에서 직접 수량/단위 읽기
                        qty_val = row[3] if len(row) > 3 else None
                        unit_val = row[4] if len(row) > 4 else None
                        
                        try:
                            qty = float(qty_val) if qty_val else 0
                        except:
                            qty = 0
                        
                        if qty <= 0:
                            continue
                        
                        unit = str(unit_val).strip() if unit_val else ""
                        
                        # 항목 객체 생성
                        item = {
                            'name': name,
                            'spec': spec,
                            'qty': qty,
                            'unit': unit,
                            'district': current_district  # None일 수도 있음
                        }

                        # 호표 참조 추출 → (name, spec) 맵에 저장 (calc_days_priority 1순위용).
                        # 계층구조 루프와 동일한 name/spec/row를 쓰므로 키 불일치가 발생하지 않는다.
                        _hopyo_num = None
                        for _v in row:
                            if isinstance(_v, str):
                                _m = re.search(r'산근\s*(\d+)\s*호표', _v)
                                if _m:
                                    _hopyo_num = int(_m.group(1))
                                    break
                        if _hopyo_num is not None:
                            item['hopyo'] = _hopyo_num
                            st.session_state["hopyo_by_item"][(name, spec)] = _hopyo_num
                        
                        # 🔥 추진공 특수 처리: #N 추진 하위 항목은 상위 "추진공"으로 합산
                        if (current_sub_sub_category and 
                            current_sub_category and 
                            "추진" in current_sub_category.get('name', '') and
                            re.match(r'^#\d+', current_sub_sub_category.get('name', ''))):
                            # #1 추진, #2 추진 등의 항목은 상위 "1) 추진공"에 합산
                            existing_item = next((i for i in current_sub_category['items'] 
                                                 if i['name'] == item['name'] and i.get('spec') == item.get('spec')), None)
                            if existing_item:
                                existing_item['qty'] = existing_item.get('qty', 0) + item.get('qty', 0)
                                msg = f"🔧 추진공 수량 합산(#N→상위): [{current_sub_category['name']}] {item['name']} ({item.get('spec', '')}) +{item.get('qty', 0)} → 총 {existing_item['qty']} {item.get('unit', '')}"
                                print(f"  {msg}")
                                with open("debug_log.txt", "a", encoding="utf-8") as f:
                                    f.write(f"{msg}\n")
                            else:
                                current_sub_category['items'].append(item)
                                msg = f"🆕 추진공 항목 추가(#N→상위): [{current_sub_category['name']}] {item['name']} ({item.get('spec', '')}) {item.get('qty', 0)} {item.get('unit', '')}"
                                print(f"  {msg}")
                                with open("debug_log.txt", "a", encoding="utf-8") as f:
                                    f.write(f"{msg}\n")
                        # sub_sub_category가 있으면 거기에 추가 (일반 케이스)
                        elif current_sub_sub_category:
                            existing_item = next((i for i in current_sub_sub_category['items']
                                                 if i['name'] == item['name'] and i.get('spec') == item.get('spec')), None)
                            if existing_item:
                                existing_item['qty'] = existing_item.get('qty', 0) + item.get('qty', 0)
                                print(f"  ✨ 수량 합산(sub_sub): {item['name']} ({item.get('spec', '')}) → {existing_item['qty']}")
                            else:
                                current_sub_sub_category['items'].append(item)
                        # sub_category에 추가
                        elif current_sub_category:
                            # 같은 name+spec 항목이 있으면 수량 합산
                            existing_item = next((i for i in current_sub_category['items'] 
                                                 if i['name'] == item['name'] and i.get('spec') == item.get('spec')), None)
                            if existing_item:
                                existing_item['qty'] = existing_item.get('qty', 0) + item.get('qty', 0)
                                # 추진공 관련은 자세히 로그
                                if "추진" in current_sub_category.get('name', ''):
                                    print(f"  🔧 추진공 수량 합산: [{current_sub_category['name']}] {item['name']} ({item.get('spec', '')}) {item.get('qty', 0)} {item.get('unit', '')} → 총 {existing_item['qty']}")
                                else:
                                    print(f"  ✨ 수량 합산: {item['name']} ({item.get('spec', '')}) → {existing_item['qty']}")
                            else:
                                current_sub_category['items'].append(item)
                                if "추진" in current_sub_category.get('name', ''):
                                    msg = f"🆕 추진공 항목 추가: [{current_sub_category['name']}] {item['name']} ({item.get('spec', '')}) {item.get('qty', 0)} {item.get('unit', '')}"
                                    print(f"  {msg}")
                                    with open("debug_log.txt", "a", encoding="utf-8") as f:
                                        f.write(f"{msg}\n")
                        else:
                            # category에 직접 추가할 때도 같은 로직
                            existing_item = next((i for i in current_category['items'] 
                                                 if i['name'] == item['name'] and i.get('spec') == item.get('spec')), None)
                            if existing_item:
                                existing_item['qty'] = existing_item.get('qty', 0) + item.get('qty', 0)
                                print(f"  ✨ 수량 합산: {item['name']} ({item.get('spec', '')}) → {existing_item['qty']}")
                            else:
                                current_category['items'].append(item)
                
                if current_category:
                    # sub_sub_category 마무리
                    if current_sub_sub_category and current_sub_category:
                        if 'sub_categories' not in current_sub_category:
                            current_sub_category['sub_categories'] = []
                        is_already_in_list = any(s is current_sub_sub_category for s in current_sub_category['sub_categories'])
                        if not is_already_in_list:
                            current_sub_category['sub_categories'].append(current_sub_sub_category)
                    
                    # sub_category 마무리
                    if current_sub_category:
                        # 같은 객체가 이미 리스트에 있는지 체크 (객체 ID 비교)
                        is_already_in_list = any(s is current_sub_category for s in current_category['sub_categories'])
                        if not is_already_in_list:
                            current_category['sub_categories'].append(current_sub_category)
                    if current_category.get('items') or current_category.get('sub_categories'):
                        hierarchy.append(current_category)
                
                if hierarchy:
                    # 중간 번호 기준 그룹핑 (1.1.X, 1.2.X, 2.1.X... 구분)
                    major_groups = {}
                    
                    for cat in hierarchy:
                        level = cat['level']
                        name = cat['name']
                        
                        # 중간 번호 추출 (1.1.X → "1.1")
                        parts = level.split('.')
                        if len(parts) >= 2:
                            major_key = f"{parts[0]}.{parts[1]}"
                        else:
                            major_key = parts[0]
                        
                        if major_key not in major_groups:
                            major_groups[major_key] = []
                        major_groups[major_key].append(cat)
                    
                    # 각 그룹 내에서 번호 순서 정렬
                    for major_key in major_groups:
                        major_groups[major_key].sort(key=lambda x: tuple(int(p) for p in x['level'].split('.')))
                    
                    # 프로그레스 완료
                    progress_bar.progress(100)
                    status_text.empty()
                    progress_bar.empty()
                    
                    st.success(f"✅ 파싱 완료! {len(major_groups)}개 공종 그룹, {sum(len(v) for v in major_groups.values())}개 주공종 인식")
                    
                    # 그룹명 정의
                    major_names = {
                        "1.1": "🏗️ 하수관로공사",
                        "1.2": "🔧 관로 부대공사",
                        "2.1": "💧 배수설비공사",
                        "2.2": "⚙️ 기계설비",
                        "3.1": "⚡ 전기공사",
                    }
                    
                    # 탭 생성
                    sorted_keys = sorted(major_groups.keys(), key=lambda x: tuple(int(p) for p in x.split('.')))
                    tab_labels = [major_names.get(key, f"📁 {key}") for key in sorted_keys]
                    
                    major_tabs = st.tabs(tab_labels)
                    
                    # 🔧 session_state 초기화
                    if 'crew_by_main' not in st.session_state:
                        st.session_state['crew_by_main'] = {}
                    
                    # 🔧 모든 카테고리 crew 기본값 미리 설정 (TAB 열기 전에!)
                    all_crew_settings = {}
                    for major_key in sorted_keys:
                        for cat in major_groups[major_key]:
                            cat_name = cat['name']
                            cat_level = cat['level']
                            cat_full = f"{cat_level} {cat_name}"
                            # 기본값 3조
                            all_crew_settings[cat_name] = st.session_state['crew_by_main'].get(cat_full, 3)
                    
                    for tab_idx, (major_key, major_tab) in enumerate(zip(sorted_keys, major_tabs)):
                        with major_tab:
                            cats_in_major = major_groups[major_key]
                            
                            st.markdown(f"### 🔧 투입조수 설정")
                            
                            if 'crew_by_main' not in st.session_state:
                                st.session_state['crew_by_main'] = {}
                            
                            cols = st.columns(min(len(cats_in_major), 4))
                            
                            for idx, cat in enumerate(cats_in_major):
                                cat_level = cat['level']
                                cat_name = cat['name']
                                cat_full = f"{cat_level} {cat_name}"
                                
                                default_crew = st.session_state['crew_by_main'].get(cat_full, 3)
                                
                                with cols[idx % len(cols)]:
                                    crew_val = st.number_input(
                                        f"{cat_full}(조)",
                                        min_value=1,
                                        max_value=30,
                                        value=default_crew,
                                        key=f"crew_{major_key.replace('.', '_')}_{idx}"
                                    )
                                    all_crew_settings[cat_name] = crew_val
                                    st.session_state['crew_by_main'][cat_full] = crew_val
                    
                    crew_settings = all_crew_settings
                    
                    st.markdown("---")
                    st.markdown("### 📊 공종별 작업일수 계산 결과")
                    
                    result_rows = []
                    
                    for cat in hierarchy:
                        cat_name = cat['name']
                        cat_level = cat['level']
                        cat_crew = crew_settings.get(cat_name, 3)
                        
                        all_cat_items = list(cat.get('items', []))
                        for sub in cat.get('sub_categories', []):
                            all_cat_items.extend(sub.get('items', []))
                            # sub_sub_categories도 포함 (3단계 계층)
                            for sub_sub in sub.get('sub_categories', []):
                                all_cat_items.extend(sub_sub.get('items', []))
                        
                        cat_total_days = sum(
                            calc_days_priority(item['name'], item.get('spec', ''), item.get('qty', 0), cat_crew)[0]
                            for item in all_cat_items
                        )
                        
                        if all_cat_items:
                            result_rows.append({
                                "level": cat_level,
                                "공종": f"{cat_level} {cat_name}",
                                "물량": f"{len(all_cat_items)}개 항목",
                                "투입조수": f"{cat_crew}조",
                                "작업일수(일)": int(cat_total_days),
                                "세부항목": all_cat_items,
                                "하위카테고리": cat.get('sub_categories', []),
                                "crew": cat_crew,
                                "major_key": '.'.join(cat_level.split('.')[:2])  # "1.1", "2.1" 등
                            })
                    
                    # 정렬
                    def sort_key(row):
                        level = row['level']
                        parts = level.split('.')
                        return tuple(int(p) for p in parts)
                    
                    result_rows_sorted = sorted(result_rows, key=sort_key)
                    
                    # ═══════════════════════════════════════════════════════
                    # 같은 공종명끼리 합산 (토공 + 토공 → 토공)
                    # ═══════════════════════════════════════════════════════
                    merged_rows = {}
                    for row in result_rows_sorted:
                        # 공종명 추출 (번호 제거: "1.1.1 토공" → "토공", "1.2.1 관로 부대공사" → "관로 부대공사")
                        parts = row['공종'].split(maxsplit=1)  # 첫 번째 공백으로만 분리
                        if len(parts) > 1 and parts[0][0].isdigit():
                            cat_name = parts[1]  # 번호 이후 전체를 공종명으로
                        else:
                            cat_name = row['공종']
                        
                        # 키를 major_key + 공종명으로 설정 (같은 이름이라도 다른 그룹은 구분)
                        merge_key = f"{row['major_key']}_{cat_name}"
                        
                        if merge_key not in merged_rows:
                            merged_rows[merge_key] = {
                                "level": row['level'],
                                "공종": cat_name,
                                "물량": 0,
                                "투입조수": row['투입조수'],
                                "작업일수(일)": 0,
                                "세부항목": [],
                                "하위카테고리": [],
                                "crew": row['crew'],
                                "major_key": row['major_key'],
                                "원본_공종들": []
                            }
                        
                        # 합산
                        merged_rows[merge_key]["작업일수(일)"] += row["작업일수(일)"]
                        merged_rows[merge_key]["세부항목"].extend(row["세부항목"])
                        merged_rows[merge_key]["하위카테고리"].extend(row["하위카테고리"])
                        merged_rows[merge_key]["원본_공종들"].append(row['공종'])
                    
                    # 물량 정보 업데이트
                    for merge_key, row in merged_rows.items():
                        total_items = len(row["세부항목"])
                        row["물량"] = f"{total_items}개 항목"
                        
                        # 공종명 정리: "1.1_관로 부대공사" → "관로 부대공사"
                        pure_name = row["공종"]  # 이미 번호 제거된 순수 공종명
                        
                        if len(row["원본_공종들"]) > 1:
                            row["공종"] = f"{pure_name} (통합)"
                        else:
                            row["공종"] = pure_name
                        
                        # major_key가 없으면 첫 번째 원본 공종에서 추출
                        if not row.get("major_key") and row.get("원본_공종들"):
                            first_gong_jong = row["원본_공종들"][0]
                            # "1.2.1 관로 부대공사" → "1.2"
                            parts = first_gong_jong.split(maxsplit=1)
                            if parts and parts[0][0].isdigit():
                                level_parts = parts[0].split('.')
                                if len(level_parts) >= 2:
                                    row["major_key"] = f"{level_parts[0]}.{level_parts[1]}"
                    
                    result_rows_merged = list(merged_rows.values())
                    max_days = max((r["작업일수(일)"] for r in result_rows_merged), default=0)
                    
                    # 그룹별로 표시
                    grouped_results = {}
                    for row in result_rows_merged:
                        major_key = row['major_key']
                        if major_key not in grouped_results:
                            grouped_results[major_key] = []
                        grouped_results[major_key].append(row)
                    
                    # 🔧 session_state에 저장 (TAB 2에서 사용)
                    st.session_state['grouped_results'] = grouped_results
                    st.session_state['group_names'] = {
                        "1.1": "🏗️ 하수관로공사",
                        "1.2": "🔧 관로 부대공사",
                        "2.1": "💧 배수설비공사",
                        "2.2": "⚙️ 기계설비",
                    }
                    st.session_state['max_days'] = max_days
                    
                    # 그룹명
                    group_names = {
                        "1.1": "🏗️ 하수관로공사",
                        "1.2": "🔧 관로 부대공사",
                        "2.1": "💧 배수설비공사",
                        "2.2": "⚙️ 기계설비",
                    }
                    
                    # 그룹별 expander
                    for major_key in sorted(grouped_results.keys(), key=lambda x: tuple(int(p) for p in x.split('.'))):
                        group_name = group_names.get(major_key, f"📁 {major_key}")
                        rows_in_group = grouped_results[major_key]
                        
                        # outer expander를 닫힌 상태로 시작 (성능 향상)
                        with st.expander(f"**{group_name}** ({len(rows_in_group)}개 공종)", expanded=False):
                            for idx, row in enumerate(rows_in_group):
                                is_max = (row["작업일수(일)"] == max_days and max_days > 0)
                                
                                with st.expander(
                                    f"{'🔴' if is_max else '▶'} **{row['공종']}** - {row['작업일수(일)']}일",
                                    expanded=False
                                ):
                                    # 하위 카테고리별 표시
                                    if row['하위카테고리']:
                                        # 같은 name의 sub_category를 하나로 합치기
                                        merged_by_name = {}
                                        for sub in row['하위카테고리']:
                                            sub_name = sub['name']
                                            if sub_name not in merged_by_name:
                                                merged_by_name[sub_name] = {
                                                    'level': sub.get('level', ''),  # 첫 번째 level 사용
                                                    'name': sub_name,
                                                    'items': [],
                                                    'sub_categories': []  # 3단계 계층
                                                }
                                            # 항목 합치기 (name+spec 기준으로 중복 제거하면서)
                                            for item in sub.get('items', []):
                                                existing = next((i for i in merged_by_name[sub_name]['items']
                                                               if i['name'] == item['name'] and i.get('spec') == item.get('spec')), None)
                                                if existing:
                                                    existing['qty'] = existing.get('qty', 0) + item.get('qty', 0)
                                                else:
                                                    merged_by_name[sub_name]['items'].append(item)
                                            # sub_categories도 합치기 (name 기준으로 중복 제거)
                                            for sub_sub in sub.get('sub_categories', []):
                                                sub_sub_name = sub_sub['name']
                                                sub_sub_level = sub_sub.get('level', '')
                                                
                                                # 같은 name+level의 sub_sub 찾기
                                                existing_sub_sub = next((s for s in merged_by_name[sub_name]['sub_categories']
                                                                       if s['name'] == sub_sub_name and s.get('level') == sub_sub_level), None)
                                                
                                                if existing_sub_sub:
                                                    # 기존 sub_sub에 항목 합치기
                                                    for item in sub_sub.get('items', []):
                                                        existing_item = next((i for i in existing_sub_sub['items']
                                                                           if i['name'] == item['name'] and i.get('spec') == item.get('spec')), None)
                                                        if existing_item:
                                                            existing_item['qty'] = existing_item.get('qty', 0) + item.get('qty', 0)
                                                        else:
                                                            existing_sub_sub['items'].append(item)
                                                else:
                                                    # 새로운 sub_sub 추가
                                                    merged_by_name[sub_name]['sub_categories'].append(sub_sub)
                                        
                                        # 합쳐진 sub_category 표시
                                        for sub_name, sub_data in merged_by_name.items():
                                            sub_level = sub_data.get('level', '')
                                            sub_items = sub_data.get('items', [])
                                            
                                            # 항목이 없으면 건너뛰기
                                            if not sub_items and not sub_data.get('sub_categories'):
                                                continue
                                            
                                            # sub_days 계산 (sub_items + sub_sub_categories)
                                            sub_days = 0
                                            for item in sub_items:
                                                d, _, _ = calc_days_priority(item['name'], item.get('spec', ''), item.get('qty', 0), row['crew'], item.get('unit', ''))
                                                sub_days += d
                                                
                                                # 추진공 디버깅 (큰 작업일수만)
                                                if "추진" in sub_name and d > 100:
                                                    print(f"⚠️ 큰 작업일수: {sub_name} - {item['name']} ({item.get('spec', '')}) qty={item.get('qty', 0)} unit={item.get('unit', '')} → {d}일")
                                            
                                            # sub_sub_categories가 있으면 그것도 포함
                                            for sub_sub in sub_data.get('sub_categories', []):
                                                for item in sub_sub.get('items', []):
                                                    d, _, _ = calc_days_priority(item['name'], item.get('spec', ''), item.get('qty', 0), row['crew'], item.get('unit', ''))
                                                    sub_days += d
                                            
                                            # 추진공인 경우 총 연장 계산 (M 단위 항목만)
                                            total_length = 0
                                            if "추진공" in sub_name or "추진 가시설공" in sub_name:
                                                for item in sub_items:
                                                    if item.get('unit') in ['M', 'm', 'M', 'ｍ']:
                                                        total_length += item.get('qty', 0)
                                            
                                            # 헤더 표시
                                            if total_length > 0:
                                                st.markdown(f"#### {sub_level} {sub_name} (총 연장: {total_length:,.1f} M, 작업일수: {sub_days}일)")
                                            else:
                                                st.markdown(f"#### {sub_level} {sub_name} ({sub_days}일)")
                                            
                                            # 🔥 sub_items 먼저 표시 (추진공의 경우 여기에 합산됨!)
                                            if sub_items:
                                                detail_items = []
                                                for item in sub_items:
                                                    d, label, method = calc_days_priority(
                                                        item['name'],
                                                        item.get('spec', ''),
                                                        item.get('qty', 0),
                                                        row['crew'],
                                                        item.get('unit', '')
                                                    )
                                                    detail_items.append({
                                                        "세부공종": item['name'],
                                                        "규격": item.get('spec', ''),
                                                        "수량": f"{item.get('qty', 0):,.1f}",
                                                        "단위": item.get('unit', ''),
                                                        "1일작업량": label,
                                                        "투입조수": row['crew'],
                                                        "작업일수": int(d),
                                                        "출처": method
                                                    })
                                                
                                                if detail_items:
                                                    st.dataframe(
                                                        pd.DataFrame(detail_items),
                                                        hide_index=True,
                                                        width="stretch"
                                                    )
                                                    
                                                    # 🔧 수동입력 UI: 매칭 안 된 항목만
                                                    unmatched_items = [
                                                        (idx, item, detail_items[idx]) 
                                                        for idx, item in enumerate(sub_items)
                                                        if detail_items[idx]["출처"] == "매칭 안 됨"
                                                    ]
                                                    
                                                    if unmatched_items:
                                                        st.markdown("---")
                                                        st.info(f"📝 매칭 안 된 항목: {len(unmatched_items)}개 - **TAB 6 '수동입력 관리'**에서 입력하세요!")
                                                        
                                                        # 🔧 session_state에 매칭 안 된 항목 수집
                                                        if "unmatched_all" not in st.session_state:
                                                            st.session_state["unmatched_all"] = {}
                                                        
                                                        # 카테고리별로 저장
                                                        cat_key = f"{major_key}_{row['공종']}"
                                                        if cat_key not in st.session_state["unmatched_all"]:
                                                            st.session_state["unmatched_all"][cat_key] = {
                                                                "major_key": major_key,
                                                                "group_name": group_name,
                                                                "category": row['공종'],
                                                                "items": []
                                                            }
                                                        
                                                        for idx_um, item_um, _ in unmatched_items:
                                                            manual_key_um = f"{item_um['name']}|{item_um.get('spec', '')}"
                                                            # 중복 방지
                                                            existing = [i for i in st.session_state["unmatched_all"][cat_key]["items"] if i["manual_key"] == manual_key_um]
                                                            if not existing:
                                                                st.session_state["unmatched_all"][cat_key]["items"].append({
                                                                    "name": item_um['name'],
                                                                    "spec": item_um.get('spec', ''),
                                                                    "qty": item_um.get('qty', 0),
                                                                    "unit": item_um.get('unit', ''),
                                                                    "manual_key": manual_key_um,
                                                                    "sub_name": sub_name
                                                                })
                                                        
                                                        # 🔧 수동입력 UI 일시 비활성화 (성능 문제로)
                                                        # TODO: 추후 별도 페이지로 분리
                                                        if False:  # 비활성화
                                                            # session_state 초기화
                                                            if "manual_rates" not in st.session_state:
                                                                st.session_state["manual_rates"] = {}
                                                            
                                                            for idx, item, detail_row in unmatched_items:
                                                                manual_key = f"{item['name']}|{item.get('spec', '')}"
                                                                
                                                                # 고유 키 생성 (카테고리명 포함으로 완전히 고유하게!)
                                                            cat_name = row['공종']
                                                            unique_key = f"{major_key}_{cat_name}_{sub_level}_{sub_name}_{manual_key}_{idx}".replace(" ", "_").replace("(", "").replace(")", "").replace("/", "_").replace("|", "_")
                                                            
                                                            col1, col2, col3, col4 = st.columns([3, 2, 1, 1])
                                                            
                                                            with col1:
                                                                st.text(f"{item['name']} ({item.get('spec', '')})")
                                                            
                                                            with col2:
                                                                # 기존 값 불러오기
                                                                existing_val = 0
                                                                existing_unit = item.get('unit', '') + "/일"
                                                                if manual_key in st.session_state["manual_rates"]:
                                                                    existing_val = st.session_state["manual_rates"][manual_key].get("daily", 0)
                                                                    existing_unit = st.session_state["manual_rates"][manual_key].get("unit", existing_unit)
                                                                
                                                                daily_rate = st.number_input(
                                                                    "1일 작업량",
                                                                    min_value=0.0,
                                                                    value=float(existing_val),
                                                                    step=0.1,
                                                                    key=f"manual_input_{unique_key}",
                                                                    label_visibility="collapsed"
                                                                )
                                                            
                                                            with col3:
                                                                unit_input = st.text_input(
                                                                    "단위",
                                                                    value=existing_unit,
                                                                    key=f"manual_unit_{unique_key}",
                                                                    label_visibility="collapsed"
                                                                )
                                                            
                                                            with col4:
                                                                if st.button("저장", key=f"manual_save_{unique_key}"):
                                                                    if daily_rate > 0:
                                                                        st.session_state["manual_rates"][manual_key] = {
                                                                            "daily": daily_rate,
                                                                            "unit": unit_input
                                                                        }
                                                                        st.success("✅ 저장됨!")
                                                                        st.rerun()
                                                            
                                                            # 계산 결과 미리보기
                                                            if daily_rate > 0:
                                                                calc_days = math.ceil(item.get('qty', 0) / (daily_rate * row['crew']))
                                                                st.caption(f"→ 예상 작업일수: **{calc_days}일** (조수: {row['crew']})")
                                            
                                            # sub_sub_categories가 있으면 추가로 표시 (3단계 계층)
                                            if sub_data.get('sub_categories'):
                                                for sub_sub in sub_data['sub_categories']:
                                                    sub_sub_level = sub_sub.get('level', '')
                                                    sub_sub_name = sub_sub['name']
                                                    sub_sub_items = sub_sub.get('items', [])
                                                    
                                                    if not sub_sub_items:
                                                        continue
                                                    
                                                    sub_sub_days = sum(
                                                        calc_days_priority(item['name'], item.get('spec', ''), item.get('qty', 0), row['crew'], item.get('unit', ''))[0]
                                                        for item in sub_sub_items
                                                    )
                                                    
                                                    # sub_sub 헤더 (들여쓰기로 계층 표현)
                                                    st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;**{sub_sub_level} {sub_sub_name}** ({sub_sub_days}일)")
                                                    
                                                    detail_items = []
                                                    for item in sub_sub_items:
                                                        d, label, method = calc_days_priority(
                                                            item['name'],
                                                            item.get('spec', ''),
                                                            item.get('qty', 0),
                                                            row['crew'],
                                                            item.get('unit', '')
                                                        )
                                                        detail_items.append({
                                                            "세부공종": item['name'],
                                                            "규격": item.get('spec', ''),
                                                            "수량": f"{item.get('qty', 0):,.1f}",
                                                            "단위": item.get('unit', ''),
                                                            "1일작업량": label,
                                                            "투입조수": row['crew'],
                                                            "작업일수": int(d),
                                                            "출처": method
                                                        })
                                                    
                                                    if detail_items:
                                                        st.dataframe(
                                                            pd.DataFrame(detail_items),
                                                            hide_index=True,
                                                            width="stretch"
                                                        )
                                            
                                            # (아래 elif 블록은 제거됨 - 위에서 이미 처리)
                                            if False:
                                                detail_items = []
                                                for item in sub_items:
                                                    d, label, method = calc_days_priority(
                                                        item['name'],
                                                        item.get('spec', ''),
                                                        item.get('qty', 0),
                                                        row['crew'],
                                                        item.get('unit', '')
                                                    )
                                                    detail_items.append({
                                                        "세부공종": item['name'],
                                                        "규격": item.get('spec', ''),
                                                        "수량": f"{item.get('qty', 0):,.1f}",
                                                        "단위": item.get('unit', ''),
                                                        "1일작업량": label,
                                                        "투입조수": row['crew'],
                                                        "작업일수": int(d),
                                                        "출처": method
                                                    })
                                                
                                                if detail_items:
                                                    st.dataframe(
                                                        pd.DataFrame(detail_items),
                                                        hide_index=True,
                                                        width="stretch"
                                                    )
                                    
                                    # 직접 항목도 있으면 표시
                                    direct_items = [item for item in row['세부항목'] if item not in sum([sub['items'] for sub in row['하위카테고리']], [])]
                                    if direct_items:
                                        detail_items = []
                                        for item in direct_items:
                                            d, label, method = calc_days_priority(
                                                item['name'],
                                                item.get('spec', ''),
                                                item.get('qty', 0),
                                                row['crew'],
                                                item.get('unit', '')
                                            )
                                            detail_items.append({
                                                "세부공종": item['name'],
                                                "규격": item.get('spec', ''),
                                                "수량": f"{item.get('qty', 0):,.1f}",
                                                "단위": item.get('unit', ''),
                                                "1일작업량": label,
                                                "투입조수": row['crew'],
                                                "작업일수": int(d),
                                                "출처": method
                                            })
                                        
                                        if detail_items:
                                            st.dataframe(
                                                pd.DataFrame(detail_items),
                                                hide_index=True,
                                                width="stretch"
                                            )
                    
                    ca, cb = st.columns(2)
                    ca.metric("🔴 주공정 (최장)", f"{max_days}일")
                    cb.metric("총 공종", f"{len(result_rows_merged)}개")
                    
                    # session_state 저장
                    st.session_state["work_result"] = {
                        "rows": result_rows_merged,
                        "hierarchy": hierarchy,
                        "crew_settings": crew_settings,
                    }
                    st.session_state["total_work_days"] = int(max_days)
                
                # ══════════════════════════════════════════════════════════════
                # 지구별 상세 섹션
                # ══════════════════════════════════════════════════════════════
                if col_info.get("districts"):
                    st.markdown("---")
                    st.markdown("## 📍 지구별 상세")
                    
                    # 지구별 데이터 그룹핑
                    district_data = {}
                    district_names = col_info["districts"]
                    
                    for item in matched:
                        district = item.get("district", "전체")
                        if district not in district_data:
                            district_data[district] = []
                        district_data[district].append(item)
                    
                    # 지구 선택
                    selected_district = st.selectbox(
                        "🏗️ 지구 선택",
                        options=list(district_data.keys()),
                        format_func=lambda x: f"{x}. {district_names.get(x, {}).get('name', x)}",
                        key="district_selector"
                    )
                    
                    selected_data = district_data[selected_district]
                    
                    st.markdown(f"### {selected_district}. {district_names.get(selected_district, {}).get('name', selected_district)}")
                    st.caption(f"총 {len(selected_data)}개 항목")
                    
                    # 공종별 그룹핑
                    group_names = {
                        "포장복구": "🛣️ 포장공",
                        "굴착공": "⛏️ 토공",
                        "관부설공": "🔧 관부설공",
                        "되메우기": "📦 되메우기",
                        "맨홀공": "🕳️ 구조물공",
                        "배수설비": "💧 배수설비",
                        "추진공": "🚇 추진공",
                        "기타": "📋 기타"
                    }
                    
                    grouped_by_type = {}
                    for item in selected_data:
                        group = item.get("group", "기타")
                        if group not in grouped_by_type:
                            grouped_by_type[group] = []
                        grouped_by_type[group].append(item)
                    
                    # 공종별 투입조수 설정
                    st.markdown("#### 🔧 투입조수 설정")
                    
                    # 해당 지구 총 공사기간 계산 및 표시 (투입조수 설정 전에)
                    # 임시로 기본 조수 3으로 계산
                    temp_total_days = 0
                    for group, items in grouped_by_type.items():
                        for item in items:
                            days, _, _ = calc_days_priority(
                                item["name"],
                                item.get("spec", ""),
                                item.get("qty", 0),
                                3  # 기본 조수
                            )
                            temp_total_days = max(temp_total_days, days)
                    
                    st.info(f"📊 **{selected_district} 지구 예상 공사기간:** {int(temp_total_days)}일 (기본 3조 기준)")
                    
                    group_crews = {}
                    
                    cols = st.columns(4)
                    for idx, (group, items) in enumerate(grouped_by_type.items()):
                        with cols[idx % 4]:
                            crew = st.number_input(
                                f"{group_names.get(group, group)}",
                                min_value=1,
                                max_value=30,
                                value=3,
                                key=f"crew_{selected_district}_{group}"
                            )
                            group_crews[group] = crew
                    
                    st.markdown("---")
                    
                    # 공종별 표 표시
                    for group in ["포장복구", "맨홀공", "굴착공", "관부설공", "되메우기", "배수설비", "추진공", "기타"]:
                        if group not in grouped_by_type:
                            continue
                        
                        items = grouped_by_type[group]
                        crew = group_crews[group]
                        
                        with st.expander(f"{group_names.get(group, group)} ({len(items)}개 항목)", expanded=(group in ["포장복구", "맨홀공", "관부설공"])):
                            display_items = []
                            for item in items:
                                days, label, method = calc_days_priority(
                                    item["name"],
                                    item.get("spec", ""),
                                    item.get("qty", 0),
                                    crew
                                )
                                display_items.append({
                                    "공종": item["name"],
                                    "규격": item.get("spec", ""),
                                    "물량": item.get("qty", 0),
                                    "단위": item.get("unit", ""),
                                    "1일작업량": label,  # "5본/일x3조" → "5본/일"로 변경 필요
                                    "조": crew,
                                    "일수": int(days),
                                    "출처": method
                                })
                            
                            if display_items:
                                display_df = pd.DataFrame(display_items)
                                
                                st.dataframe(
                                    display_df,
                                    width="stretch",
                                    height=400,
                                    column_config={
                                        "공종": st.column_config.TextColumn("공종", width="large"),
                                        "규격": st.column_config.TextColumn("규격", width="large"),
                                        "물량": st.column_config.NumberColumn("물량", width="medium", format="%.1f"),
                                        "단위": st.column_config.TextColumn("단위", width="small"),
                                        "1일작업량": st.column_config.TextColumn("1일작업량", width="medium"),
                                        "조": st.column_config.NumberColumn("조", width="small"),
                                        "일수": st.column_config.NumberColumn("일수", width="small"),
                                        "출처": st.column_config.TextColumn("출처", width="medium"),
                                    }
                                )
                                
                                total_days = sum(item["일수"] for item in display_items)
                                st.metric(f"{group_names.get(group, group)} 총 작업일수", f"{total_days}일")
                    
        except Exception as e:
            st.error(f"파싱 실패: {e}")
            import traceback
            st.code(traceback.format_exc())
    else:
        st.info("도급 설계내역서 엑셀을 업로드해주세요.")

# ══════════════════════════════════════════════════════════════
# TAB 1
# ══════════════════════════════════════════════════════════════
with tab1:
    st.subheader("📋 공기산정 요약")
    
    # session_state에서 지역 가져오기 (비작업일수 탭에서 설정)
    selected_region = st.session_state.get("selected_region", "서울")
    
    # 지역이 설정되지 않은 경우 안내
    if "selected_region" not in st.session_state:
        st.warning("⚠️ 먼저 **'비작업일수 계산기'** 탭에서 공사 지역을 선택해주세요!")
    
    st.markdown("### 📊 공사 정보 요약")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("📍 공사 지역", selected_region)
    
    with col2:
        total_work_days = st.session_state.get("total_work_days", 0)
        st.metric("💼 총 순작업일수", f"{total_work_days}일")
    
    with col3:
        # 비작업일수 결과 표시 (있는 경우)
        if "weather_result" in st.session_state:
            non_work = st.session_state["weather_result"].get("non_work_days", 0)
            st.metric("🚫 비작업일수", f"{non_work}일")
        else:
            st.metric("🚫 비작업일수", "미계산")
    
    st.markdown("---")
    
    if "work_result" in st.session_state:
        st.success("✅ 엑셀 인식 탭에서 계산 완료!")
        
        # 최종 결과 (비작업일수 계산 완료된 경우)
        if "weather_result" in st.session_state:
            result = st.session_state["weather_result"]
            
            st.markdown("### 🎯 최종 공기산정 결과")
            
            col_a, col_b, col_c, col_d = st.columns(4)
            col_a.metric("📅 총 공사기간", f"{result['total_days']}일")
            col_b.metric("💼 순작업일수", f"{result['work_days']}일")
            col_c.metric("🚫 비작업일수", f"{result['non_work_days']}일")
            col_d.metric("📍 적용 지역", result['region'])
            
            st.info(f"""
            **📍 {result['region']} 지역 공기산정 결과**
            - 착공일: {result['start_date'].strftime('%Y년 %m월 %d일')}
            - 준공일: {result['end_date'].strftime('%Y년 %m월 %d일')}
            - 총 공사기간: **{result['total_days']}일**
            
            **적용된 비작업일 조건:**
            - {'✅' if result['include_rain'] else '❌'} 강우일
            - {'✅' if result['include_cold'] else '❌'} 한랭일
            - {'✅' if result['include_hot'] else '❌'} 폭염일
            """)
        else:
            st.info("👉 **'비작업일수 계산기'** 탭에서 비작업일수를 계산하면 최종 공기산정 결과가 표시됩니다!")
    else:
        st.warning("⚠️ **'엑셀 내역서 인식'** 탭에서 엑셀을 먼저 업로드해주세요.")

# ══════════════════════════════════════════════════════════════
# TAB 3
# ══════════════════════════════════════════════════════════════
with tab3:
    st.subheader("주요공종 CP 분석")
    work_result = st.session_state.get("work_result")
    if work_result:
        result_rows = work_result["rows"]
        
        display_data = []
        for r in result_rows:
            display_data.append({
                "공종": r["공종"],
                "물량": r["물량"],
                "투입조수": r["투입조수"],
                "작업일수(일)": r["작업일수(일)"]
            })
        
        df_cp = pd.DataFrame(display_data)
        df_cp = df_cp[df_cp["작업일수(일)"] > 0].copy()
        max_days = df_cp["작업일수(일)"].max() if len(df_cp) > 0 else 0

        def hl_cp(row):
            if row["작업일수(일)"] == max_days:
                return ["background-color:#3d0000;color:#ff6b6b"] * len(row)
            return [""] * len(row)

        st.dataframe(df_cp.style.apply(hl_cp, axis=1), hide_index=True, width="stretch")
        
        fig_bar = px.bar(df_cp, x="작업일수(일)", y="공종", orientation="h", text="작업일수(일)",
                         color="작업일수(일)", color_continuous_scale=["#27AE60","#F39C12","#E74C3C"])
        fig_bar.update_layout(height=350, showlegend=False, yaxis=dict(autorange="reversed"))
        st.plotly_chart(fig_bar, width="stretch")
    else:
        st.warning("TAB 2에서 엑셀을 먼저 업로드해주세요.")

# ══════════════════════════════════════════════════════════════
# TAB 4
# ══════════════════════════════════════════════════════════════
with tab4:
    st.subheader("🌧 비작업일수 계산기")
    st.caption("지역별 기상 데이터를 기반으로 비작업일수를 계산합니다")
    
    # ──────────────────────────────────
    # 1. 지역 선택
    # ──────────────────────────────────
    st.markdown("### 🌍 공사 지역 선택")
    
    col_r1, col_r2 = st.columns([2, 1])
    with col_r1:
        # 17개 지역
        default_region = st.session_state.get("selected_region", "서울")
        if default_region not in REGIONS:
            default_region = "서울"
        
        selected_region = st.selectbox(
            "공사 지역",
            options=REGIONS,
            index=REGIONS.index(default_region),
            help="지역별 기상 데이터 (강우/한파/폭염) 적용"
        )
        st.session_state["selected_region"] = selected_region
    
    with col_r2:
        st.metric("선택 지역", f"📍 {selected_region}")
    
    st.markdown("---")
    
    # ──────────────────────────────────
    # 2. 비작업일 조건
    # ──────────────────────────────────
    st.markdown("### ⚙️ 비작업일 조건 설정")
    st.caption("📖 가이드라인 공식: 비작업일수 = 기상조건(A) + 법정공휴일(B) - 중복일수(C)")
    
    st.markdown("**🌤️ 기상조건 (A)**")
    col1, col2, col3 = st.columns(3)
    with col1:
        include_rain = st.checkbox("🌧️ 강우일 포함", value=True, help="월별 평균 강우일수")
    with col2:
        include_cold = st.checkbox("❄️ 한랭일 포함", value=True, help="일 최저기온 -10°C 이하")
    with col3:
        include_hot = st.checkbox("🔥 폭염일 포함", value=True, help="일 최고기온 33°C 이상")
    
    st.markdown("**📅 법정공휴일 (B)**")
    col_h1, col_h2 = st.columns(2)
    with col_h1:
        include_holidays = st.checkbox(
            "📅 법정공휴일 포함",
            value=True,
            help="가이드라인 부록1: 일요일(52일) + 명절 + 국경일 + 기타 + 대체공휴일"
        )
    with col_h2:
        min_weekly_rest = st.checkbox(
            "⚖️ 주 40시간 근무제 보장",
            value=True,
            help="월별 비작업일수가 주 40시간 근무제 일수보다 작으면 보정"
        )
    
    st.markdown("---")
    
    # ──────────────────────────────────
    # 3. 공사 기간 설정
    # ──────────────────────────────────
    st.markdown("### 📅 공사 기간 설정")
    
    col_a, col_b = st.columns(2)
    with col_a:
        start_date = st.date_input(
            "착공일",
            value=st.session_state.get("start_date", datetime.now().date()),
            key="weather_start_date"
        )
        st.session_state["start_date"] = start_date
    
    with col_b:
        # TAB 2에서 계산된 순작업일수 자동 입력
        default_work_days = st.session_state.get("total_work_days", 100)
        # max_value를 동적으로 설정 (값이 크면 max도 자동으로 늘림)
        max_val = max(10000, default_work_days + 1000)
        work_days = st.number_input(
            "순작업일수",
            min_value=1,
            max_value=max_val,
            value=default_work_days,
            key="weather_work_days",
            help="TAB '엑셀 내역서 인식'에서 자동 계산된 값"
        )
        st.session_state["work_days_input"] = work_days
    
    st.markdown("---")
    
    # ──────────────────────────────────
    # 4. 계산 버튼
    # ──────────────────────────────────
    if st.button("📊 비작업일수 계산", type="primary", width="stretch"):
        from datetime import datetime as dt
        
        # datetime 변환
        start_dt = dt.combine(start_date, dt.min.time())
        
        # 종료일 추정: 순작업일수 * 1.5
        rough_end_date = start_dt + timedelta(days=int(work_days * 1.5))
        
        # 반복 계산: 정확한 종료일 찾기
        for _ in range(5):
            # 1. 기상조건 비작업일수 (A)
            weather_days = get_total_non_work_days(
                selected_region,
                start_dt,
                rough_end_date,
                check_rain=include_rain,
                check_cold=include_cold,
                check_hot=include_hot
            )
            
            if isinstance(weather_days, dict):
                weather_days = weather_days.get("total", 0)
            
            # 2. 가이드라인 공식 적용 (A + B - C)
            result = get_total_non_work_days_with_holidays(
                weather_days,
                start_dt,
                rough_end_date,
                include_holidays=include_holidays,
                min_weekly_rest=min_weekly_rest
            )
            
            non_work_days = result["total"]
            
            # 총 공사기간 = 순작업일수 + 비작업일수
            calculated_end = start_dt + timedelta(days=int(work_days + non_work_days - 1))
            
            # 수렴 체크
            if abs((rough_end_date - calculated_end).days) <= 1:
                break
            rough_end_date = calculated_end
        
        completion_date = calculated_end
        total_days = (completion_date - start_dt).days + 1
        
        # 결과 저장
        st.session_state["weather_result"] = {
            "region": selected_region,
            "start_date": start_dt,
            "end_date": completion_date,
            "total_days": total_days,
            "work_days": work_days,
            "non_work_days": non_work_days,
            "weather_days": result["weather"],
            "holiday_days": result["holidays"],
            "overlap_days": result["overlap"],
            "formula": result["formula"],
            "include_rain": include_rain,
            "include_cold": include_cold,
            "include_hot": include_hot,
            "include_holidays": include_holidays,
            "min_weekly_rest": min_weekly_rest,
        }
        
        st.success(f"✅ 준공일: **{completion_date.strftime('%Y년 %m월 %d일')}**")
        # 공기산정 탭(tab1)이 이 탭보다 먼저 실행되므로, 방금 저장한 weather_result를
        # 같은 실행에서 못 읽는다. 즉시 rerun 해서 모든 탭이 새 값으로 다시 렌더되게 한다.
        st.rerun()
    
    # ──────────────────────────────────
    # 5. 결과 표시
    # ──────────────────────────────────
    if "weather_result" in st.session_state:
        result = st.session_state["weather_result"]
        
        st.markdown("### 📊 계산 결과")
        
        # 메인 메트릭
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        col_m1.metric("📍 지역", result["region"])
        col_m2.metric("📅 총 공사기간", f"{result['total_days']}일")
        col_m3.metric("💼 순작업일수", f"{result['work_days']}일")
        col_m4.metric("🚫 비작업일수", f"{result['non_work_days']}일")
        
        # 비작업일수 세부 (A + B - C)
        st.markdown("#### 🧮 비작업일수 세부 (가이드라인 공식)")
        col_a, col_b, col_c, col_d = st.columns(4)
        col_a.metric("🌤️ 기상조건 (A)", f"{result.get('weather_days', 0)}일")
        col_b.metric("📅 법정공휴일 (B)", f"{result.get('holiday_days', 0)}일")
        col_c.metric("⚠️ 중복 (C)", f"{result.get('overlap_days', 0)}일")
        col_d.metric("📌 공식", f"A+B-C = {result.get('non_work_days', 0)}일")
        
        st.caption(f"📐 계산식: {result.get('formula', '')}")
        
        # 적용 조건
        st.info(f"""
        **적용된 조건:**
        - {'✅' if result['include_rain'] else '❌'} 강우일
        - {'✅' if result['include_cold'] else '❌'} 한랭일 (-10°C 이하)
        - {'✅' if result['include_hot'] else '❌'} 폭염일 (33°C 이상)
        - {'✅' if result.get('include_holidays', False) else '❌'} 법정공휴일
        - {'✅' if result.get('min_weekly_rest', False) else '❌'} 주 40시간 근무제 보장
        - **기간**: {result['start_date'].strftime('%Y-%m-%d')} ~ {result['end_date'].strftime('%Y-%m-%d')}
        """)
        
        # 월별 통합 상세 표
        try:
            from datetime import datetime as dt
            import pandas as pd
            
            monthly_weather = get_monthly_breakdown(
                result["region"],
                result["start_date"],
                result["end_date"],
                check_rain=result["include_rain"],
                check_cold=result["include_cold"],
                check_hot=result["include_hot"]
            )
            
            monthly_holidays = get_holiday_breakdown_monthly(
                result["start_date"],
                result["end_date"]
            )
            
            if monthly_weather:
                st.markdown("### 📅 월별 비작업일수 상세")
                st.caption("📖 각 월의 기상조건 + 공휴일 + 중복일수 + 최종 비작업일수")
                
                # 월별 데이터 통합
                holiday_dict = {h["월"]: h["법정공휴일"] for h in monthly_holidays} if monthly_holidays else {}
                
                monthly_data = []
                from calendar import monthrange
                
                for m in monthly_weather:
                    month_str = m["month"]
                    year, mon = map(int, month_str.split("-"))
                    cal_days = monthrange(year, mon)[1]
                    
                    rain = m.get("rain", 0)
                    cold = m.get("cold", 0)
                    hot = m.get("hot", 0)
                    weather_total = rain + cold + hot  # A
                    
                    holidays = holiday_dict.get(month_str, 0) if result.get("include_holidays", False) else 0  # B
                    
                    # 중복일수 (C = A × B ÷ 달력일수)
                    overlap = round(weather_total * holidays / cal_days) if cal_days > 0 else 0
                    
                    # 월별 비작업일수
                    month_non_work = round(weather_total + holidays - overlap)
                    
                    # 주 40시간 근무제 보장
                    weeks = cal_days / 7
                    min_rest = round(weeks)
                    if result.get("min_weekly_rest", False) and month_non_work < min_rest:
                        month_non_work = min_rest
                    
                    monthly_data.append({
                        "월": month_str,
                        "🌧️ 강우": f"{rain:.1f}",
                        "❄️ 한랭": f"{cold:.1f}",
                        "🔥 폭염": f"{hot:.1f}",
                        "🌤️ 기상(A)": f"{weather_total:.1f}",
                        "📅 공휴일(B)": f"{holidays}",
                        "⚠️ 중복(C)": f"{overlap}",
                        "📊 비작업": f"{month_non_work}",
                        "📆 달력일": f"{cal_days}",
                    })
                
                df_monthly = pd.DataFrame(monthly_data)
                st.dataframe(df_monthly, hide_index=True, width="stretch")
                
                # ──────────────────────────────────
                # 항목별 분석
                # ──────────────────────────────────
                st.markdown("### 📈 항목별 비작업일수 분석")
                
                # 합계 계산
                total_rain = sum(m.get("rain", 0) for m in monthly_weather)
                total_cold = sum(m.get("cold", 0) for m in monthly_weather)
                total_hot = sum(m.get("hot", 0) for m in monthly_weather)
                total_weather = total_rain + total_cold + total_hot
                total_holiday = sum(holiday_dict.values())
                
                # 막대 차트
                chart_data = {
                    "항목": ["🌧️ 강우일", "❄️ 한랭일", "🔥 폭염일", "📅 법정공휴일", "⚠️ 중복일수"],
                    "일수": [total_rain, total_cold, total_hot, total_holiday, result.get('overlap_days', 0)]
                }
                df_chart = pd.DataFrame(chart_data)
                
                col_chart1, col_chart2 = st.columns([2, 1])
                
                with col_chart1:
                    st.bar_chart(df_chart.set_index("항목"))
                
                with col_chart2:
                    st.markdown("**📊 합계**")
                    st.metric("🌧️ 강우일", f"{total_rain:.1f}일")
                    st.metric("❄️ 한랭일", f"{total_cold:.1f}일")
                    st.metric("🔥 폭염일", f"{total_hot:.1f}일")
                    st.metric("📅 법정공휴일", f"{total_holiday}일")
                
                # 항목별 expander
                with st.expander("🌧️ 강우일 월별 상세", expanded=False):
                    rain_data = [{"월": m["month"], "강우일수": f"{m.get('rain', 0):.1f}일"} for m in monthly_weather]
                    st.dataframe(pd.DataFrame(rain_data), hide_index=True, width="stretch")
                    st.caption(f"📌 {result['region']} 지역의 월별 평균 강우일수 (기상청 기준)")
                
                with st.expander("❄️ 한랭일 월별 상세", expanded=False):
                    cold_data = [{"월": m["month"], "한랭일수": f"{m.get('cold', 0):.1f}일"} for m in monthly_weather]
                    st.dataframe(pd.DataFrame(cold_data), hide_index=True, width="stretch")
                    st.caption(f"📌 일 최저기온 -10°C 이하 기준 (한랭일 평균)")
                
                with st.expander("🔥 폭염일 월별 상세", expanded=False):
                    hot_data = [{"월": m["month"], "폭염일수": f"{m.get('hot', 0):.1f}일"} for m in monthly_weather]
                    st.dataframe(pd.DataFrame(hot_data), hide_index=True, width="stretch")
                    st.caption(f"📌 일 최고기온 33°C 이상 기준 (폭염일 평균)")
                
                with st.expander("📅 법정공휴일 월별 상세", expanded=False):
                    if monthly_holidays:
                        holiday_detail = [{"월": h["월"], "공휴일수": f"{h['법정공휴일']}일"} for h in monthly_holidays]
                        st.dataframe(pd.DataFrame(holiday_detail), hide_index=True, width="stretch")
                        st.caption("📌 부록1: 일요일(52일) + 명절 + 국경일 + 기타 공휴일 + 대체공휴일")
                    else:
                        st.info("법정공휴일이 포함되지 않았습니다.")
                
                # 가이드라인 공식 설명
                with st.expander("📐 계산 공식 설명", expanded=False):
                    st.markdown(f"""
                    ### 가이드라인 19페이지 공식
                    
                    **비작업일수 = A + B - C**
                    
                    | 항목 | 내용 | 값 |
                    |------|------|-----|
                    | **A** | 기상조건 비작업일수 (강우 + 한랭 + 폭염) | {result.get('weather_days', 0)}일 |
                    | **B** | 법정 공휴일수 (부록1 기준) | {result.get('holiday_days', 0)}일 |
                    | **C** | 중복일수 = A × B ÷ 달력일수 (소수점 반올림) | {result.get('overlap_days', 0)}일 |
                    | **계** | A + B - C | **{result.get('non_work_days', 0)}일** |
                    
                    ### 주 40시간 근무제 보장
                    - 월별 비작업일수가 주 40시간 근무제 일수보다 작을 경우, 보정 적용
                    - 일반적으로 주 1일 휴식 보장 (월 4~5일)
                    
                    ### 데이터 출처
                    - **기상 데이터**: 기상청 평년값 (1991-2020)
                    - **법정공휴일**: 「관공서의 공휴일에 관한 규정」 (부록1, 2026-2035)
                    """)
        except Exception as e:
            st.error(f"월별 상세 정보 표시 오류: {e}")
            import traceback
            st.code(traceback.format_exc())
    
    # ──────────────────────────────────
    # 6. 지역 기상 정보 미리보기
    # ──────────────────────────────────
    with st.expander(f"📊 {selected_region} 지역 연간 기상 통계", expanded=False):
        if selected_region in RAIN_DAYS:
            import pandas as pd
            
            months = list(range(1, 13))
            data = {
                "월": [f"{m}월" for m in months],
                "🌧️ 강우일": [RAIN_DAYS[selected_region].get(m, 0) for m in months],
                "❄️ 한랭일": [COLD_DAYS[selected_region].get(m, 0) for m in months],
                "🔥 폭염일": [HOT_DAYS[selected_region].get(m, 0) for m in months],
            }
            df_stats = pd.DataFrame(data)
            df_stats["합계"] = df_stats["🌧️ 강우일"] + df_stats["❄️ 한랭일"] + df_stats["🔥 폭염일"]
            
            st.dataframe(df_stats, hide_index=True, width="stretch")
            
            # 연간 합계
            annual_rain = sum(RAIN_DAYS[selected_region].values())
            annual_cold = sum(COLD_DAYS[selected_region].values())
            annual_hot = sum(HOT_DAYS[selected_region].values())
            
            col_s1, col_s2, col_s3, col_s4 = st.columns(4)
            col_s1.metric("연간 강우일", f"{annual_rain:.1f}일")
            col_s2.metric("연간 한랭일", f"{annual_cold:.1f}일")
            col_s3.metric("연간 폭염일", f"{annual_hot:.1f}일")
            col_s4.metric("연간 총합", f"{annual_rain + annual_cold + annual_hot:.1f}일")

# ══════════════════════════════════════════════════════════════
# TAB 5
# ══════════════════════════════════════════════════════════════
with tab5:
    st.subheader("📄 공기산정 보고서")
    
    if "work_result" not in st.session_state:
        st.warning("TAB 2에서 엑셀을 먼저 업로드해주세요.")
    else:
        st.markdown("### 📊 보고서 생성")
        
        col1, col2 = st.columns(2)
        
        with col1:
            project_name = st.text_input("공사명", value="상하수도 관로공사", key="project_name")
        with col2:
            client_name = st.text_input("발주처", value="", key="client_name")
        
        col3, col4 = st.columns(2)
        with col3:
            start_date_input = st.date_input("착공일", datetime.now().date(), key="report_start_date")
        with col4:
            total_work_days = st.session_state.get('total_work_days', 0)
            st.metric("순작업일수", f"{total_work_days}일")
        
        st.markdown("---")
        
        if st.button("📥 엑셀 보고서 생성", type="primary", width="stretch"):
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from io import BytesIO
                
                # 새 워크북 생성
                wb = Workbook()
                
                # ═══════════════════════════════════════════════════════
                # Sheet 1: 표지
                # ═══════════════════════════════════════════════════════
                ws_cover = wb.active
                ws_cover.title = "표지"
                
                ws_cover['A5'] = project_name
                ws_cover['A5'].font = Font(size=18, bold=True)
                ws_cover['A5'].alignment = Alignment(horizontal='center')
                
                ws_cover['A7'] = "공사기간 산정 검토서"
                ws_cover['A7'].font = Font(size=16, bold=True)
                ws_cover['A7'].alignment = Alignment(horizontal='center')
                
                ws_cover['A10'] = f"작성일: {datetime.now().strftime('%Y년 %m월 %d일')}"
                ws_cover['A10'].alignment = Alignment(horizontal='center')
                
                # ═══════════════════════════════════════════════════════
                # Sheet 2: 공사기간 산정
                # ═══════════════════════════════════════════════════════
                ws_calc = wb.create_sheet("2. 공사기간 산정")
                
                ws_calc['A1'] = "2. 공사기간 산정"
                ws_calc['A1'].font = Font(size=14, bold=True)
                
                ws_calc['A3'] = "2.1 작업일수"
                ws_calc['A3'].font = Font(size=12, bold=True)
                
                # 공종별 작업일수 표
                headers = ["공종", "물량", "투입조수", "작업일수(일)"]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_calc.cell(row=5, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                work_result = st.session_state["work_result"]
                rows = work_result["rows"]
                
                for row_idx, row_data in enumerate(rows, 6):
                    ws_calc.cell(row=row_idx, column=1, value=row_data["공종"])
                    ws_calc.cell(row=row_idx, column=2, value=row_data["물량"])
                    ws_calc.cell(row=row_idx, column=3, value=row_data["투입조수"])
                    ws_calc.cell(row=row_idx, column=4, value=row_data["작업일수(일)"])
                    
                    for col in range(1, 5):
                        cell = ws_calc.cell(row=row_idx, column=col)
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                
                # 합계
                total_row = 6 + len(rows)
                ws_calc.cell(row=total_row, column=1, value="합계")
                ws_calc.cell(row=total_row, column=1).font = Font(bold=True)
                ws_calc.cell(row=total_row, column=4, value=total_work_days)
                ws_calc.cell(row=total_row, column=4).font = Font(bold=True)
                
                # 주공정 표시
                max_days = max((r["작업일수(일)"] for r in rows), default=0)
                ws_calc[f'A{total_row+2}'] = f"🔴 주공정 (Critical Path): {max_days}일"
                ws_calc[f'A{total_row+2}'].font = Font(bold=True, color="C00000")
                
                # ═══════════════════════════════════════════════════════
                # Sheet 3: 부록1. 작업일수 산정 (계층 구조)
                # ═══════════════════════════════════════════════════════
                ws_appendix = wb.create_sheet("부록1. 작업일수 산정근거")
                
                ws_appendix['A1'] = "◈ 부록1. 작업일수 산정근거"
                ws_appendix['A1'].font = Font(size=14, bold=True)
                
                # 헤더 행
                headers_detail = ["공종명", "규격", "수량", "단위", "1일 작업량", "조", "작업시간(HR)", "적업일수(일)", "비 고"]
                ws_appendix.merge_cells('A3:A3')
                for col_idx, header in enumerate(headers_detail, 1):
                    cell = ws_appendix.cell(row=3, column=col_idx, value=header)
                    cell.font = Font(bold=True, size=10)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                current_row = 4
                
                # hierarchy 데이터로 계층 구조 생성
                hierarchy = work_result.get("hierarchy", [])
                crew_settings = work_result.get("crew_settings", {})
                
                # 전체 합계 계산
                total_all_days = sum(
                    sum(calc_days_priority(item['name'], item.get('spec', ''), item.get('qty', 0), 
                        crew_settings.get(cat['name'], 3))[0]
                        for item in (cat.get('items', []) + 
                                    sum([sub.get('items', []) for sub in cat.get('sub_categories', [])], [])))
                    for cat in hierarchy
                )
                
                # 지구명 (있으면)
                district_name = "전체"
                if hierarchy:
                    ws_appendix.cell(row=current_row, column=1, value=f"■ {district_name}").font = Font(bold=True, size=11)
                    ws_appendix.cell(row=current_row, column=1).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    current_row += 1
                
                # 대분류별 처리
                for cat_idx, cat in enumerate(hierarchy, 1):
                    cat_name = cat['name']
                    cat_level = cat['level']
                    cat_crew = crew_settings.get(cat_name, 3)
                    
                    # 대분류 전체 일수 계산
                    all_cat_items = list(cat.get('items', []))
                    for sub in cat.get('sub_categories', []):
                        all_cat_items.extend(sub.get('items', []))
                    
                    cat_total_days = sum(
                        calc_days_priority(item['name'], item.get('spec', ''), item.get('qty', 0), cat_crew)[0]
                        for item in all_cat_items
                    )
                    
                    # 대분류 헤더 (1. 우수관로)
                    cell = ws_appendix.cell(row=current_row, column=1, value=f"{cat_idx}.{cat_name}")
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    
                    ws_appendix.cell(row=current_row, column=4, value="식").alignment = Alignment(horizontal='center')
                    ws_appendix.cell(row=current_row, column=8, value=int(cat_total_days)).font = Font(bold=True)
                    current_row += 1
                    
                    # 하위 카테고리 처리
                    sub_categories = cat.get('sub_categories', [])
                    
                    if sub_categories:
                        for sub_idx, sub in enumerate(sub_categories, 1):
                            sub_name = sub['name']
                            sub_items = sub.get('items', [])
                            
                            sub_total_days = sum(
                                calc_days_priority(item['name'], item.get('spec', ''), item.get('qty', 0), cat_crew)[0]
                                for item in sub_items
                            )
                            
                            # 중분류 헤더 (1.1 토공)
                            indent = "  "
                            cell = ws_appendix.cell(row=current_row, column=1, value=f"{indent}{cat_idx}.{sub_idx} {sub_name}")
                            cell.font = Font(bold=True)
                            ws_appendix.cell(row=current_row, column=8, value=int(sub_total_days)).font = Font(bold=True)
                            current_row += 1
                            
                            # 세부 항목들
                            for item_idx, item in enumerate(sub_items, 1):
                                days, label, method = calc_days_priority(
                                    item['name'],
                                    item.get('spec', ''),
                                    item.get('qty', 0),
                                    cat_crew
                                )
                                
                                indent2 = "    "
                                ws_appendix.cell(row=current_row, column=1, value=f"{indent2}{item_idx}) {item['name']}")
                                ws_appendix.cell(row=current_row, column=2, value=item.get('spec', ''))
                                ws_appendix.cell(row=current_row, column=3, value=item.get('qty', 0))
                                ws_appendix.cell(row=current_row, column=4, value=item.get('unit', ''))
                                ws_appendix.cell(row=current_row, column=5, value=label)
                                ws_appendix.cell(row=current_row, column=6, value=cat_crew)
                                ws_appendix.cell(row=current_row, column=7, value=8)  # 작업시간
                                ws_appendix.cell(row=current_row, column=8, value=int(days))
                                ws_appendix.cell(row=current_row, column=9, value=method)
                                
                                current_row += 1
                    
                    # 직접 항목이 있으면
                    direct_items = cat.get('items', [])
                    if direct_items:
                        for item_idx, item in enumerate(direct_items, 1):
                            days, label, method = calc_days_priority(
                                item['name'],
                                item.get('spec', ''),
                                item.get('qty', 0),
                                cat_crew
                            )
                            
                            indent = "  "
                            ws_appendix.cell(row=current_row, column=1, value=f"{indent}{item_idx}) {item['name']}")
                            ws_appendix.cell(row=current_row, column=2, value=item.get('spec', ''))
                            ws_appendix.cell(row=current_row, column=3, value=item.get('qty', 0))
                            ws_appendix.cell(row=current_row, column=4, value=item.get('unit', ''))
                            ws_appendix.cell(row=current_row, column=5, value=label)
                            ws_appendix.cell(row=current_row, column=6, value=cat_crew)
                            ws_appendix.cell(row=current_row, column=7, value=8)
                            ws_appendix.cell(row=current_row, column=8, value=int(days))
                            ws_appendix.cell(row=current_row, column=9, value=method)
                            
                            current_row += 1
                    
                    current_row += 1  # 대분류 간 여백
                
                # 컬럼 너비 조정
                ws_calc.column_dimensions['A'].width = 40
                ws_calc.column_dimensions['B'].width = 15
                ws_calc.column_dimensions['C'].width = 15
                ws_calc.column_dimensions['D'].width = 15
                
                ws_appendix.column_dimensions['A'].width = 50
                ws_appendix.column_dimensions['B'].width = 35
                ws_appendix.column_dimensions['C'].width = 12
                ws_appendix.column_dimensions['D'].width = 10
                ws_appendix.column_dimensions['E'].width = 25
                ws_appendix.column_dimensions['F'].width = 8
                ws_appendix.column_dimensions['G'].width = 15
                ws_appendix.column_dimensions['H'].width = 15
                ws_appendix.column_dimensions['I'].width = 15
                
                # BytesIO로 저장
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                # 다운로드 버튼
                st.success("✅ 보고서가 생성되었습니다!")
                st.download_button(
                    label="📥 엑셀 다운로드",
                    data=excel_buffer,
                    file_name=f"공사기간_산정_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width="stretch"
                )
                
            except Exception as e:
                st.error(f"보고서 생성 실패: {e}")
                import traceback
                st.code(traceback.format_exc())

# ══════════════════════════════════════════════════════════════
# TAB 6: 수동입력 관리
# ══════════════════════════════════════════════════════════════
with tab6:
    st.subheader("📝 수동입력 관리")
    st.caption("매칭 안 된 항목들에 대해 1일 작업량을 직접 입력하세요")
    
    # session_state 초기화
    if "manual_rates" not in st.session_state:
        st.session_state["manual_rates"] = {}
    
    if "unmatched_all" not in st.session_state or not st.session_state["unmatched_all"]:
        st.info("📂 먼저 TAB 2에서 엑셀 파일을 업로드하세요!")
    else:
        unmatched_all = st.session_state["unmatched_all"]
        
        # 그룹별로 정리
        group_data = {
            "1.1": {"name": "🏗️ 하수관로공사", "categories": []},
            "1.2": {"name": "🔧 관로 부대공사", "categories": []},
            "2.1": {"name": "💧 배수설비공사", "categories": []},
            "2.2": {"name": "⚙️ 기계설비", "categories": []},
        }
        
        for cat_key, cat_data in unmatched_all.items():
            mk = cat_data["major_key"]
            if mk in group_data:
                group_data[mk]["categories"].append(cat_data)
        
        # 전체 통계
        total_unmatched = sum(
            sum(len(cat["items"]) for cat in gd["categories"])
            for gd in group_data.values()
        )
        saved_count = len(st.session_state["manual_rates"])
        
        col_a, col_b, col_c = st.columns(3)
        with col_a:
            st.metric("전체 매칭 안 된 항목", f"{total_unmatched}개")
        with col_b:
            st.metric("저장된 항목", f"{saved_count}개", delta=f"{total_unmatched - saved_count}개 남음")
        with col_c:
            if st.button("🗑️ 모든 저장 초기화", width="stretch"):
                st.session_state["manual_rates"] = {}
                st.rerun()
        
        st.markdown("---")
        
        # 하위 탭 (4개 그룹)
        sub_tab_labels = []
        sub_tab_keys = []
        for mk, gd in group_data.items():
            count = sum(len(cat["items"]) for cat in gd["categories"])
            if count > 0:
                sub_tab_labels.append(f"{gd['name']} ({count})")
                sub_tab_keys.append(mk)
        
        if not sub_tab_labels:
            st.success("🎉 모든 항목이 매칭되었습니다!")
        else:
            sub_tabs = st.tabs(sub_tab_labels)
            
            # 🚀 @st.fragment로 페이지 변경 시 전체 재실행 방지
            @st.fragment
            def render_manual_input_page(mk, gd):
                # 모든 항목을 평탄화
                all_items = []
                for cat in gd["categories"]:
                    for item in cat["items"]:
                        all_items.append({
                            **item,
                            "category": cat["category"]
                        })
                
                if not all_items:
                    st.info("매칭 안 된 항목이 없습니다.")
                    return
                
                # 페이지네이션
                items_per_page = 20
                total_pages = (len(all_items) + items_per_page - 1) // items_per_page
                
                page_key = f"page_{mk}"
                if page_key not in st.session_state:
                    st.session_state[page_key] = 1
                
                # 페이지 선택
                col_p1, col_p2, col_p3 = st.columns([1, 2, 1])
                with col_p1:
                    if st.button("◀ 이전", key=f"prev_{mk}", disabled=(st.session_state[page_key] <= 1)):
                        st.session_state[page_key] -= 1
                        st.rerun(scope="fragment")
                with col_p2:
                    st.markdown(f"<div style='text-align: center; padding: 8px;'>페이지 {st.session_state[page_key]} / {total_pages}</div>", unsafe_allow_html=True)
                with col_p3:
                    if st.button("다음 ▶", key=f"next_{mk}", disabled=(st.session_state[page_key] >= total_pages)):
                        st.session_state[page_key] += 1
                        st.rerun(scope="fragment")
                
                # 현재 페이지 항목
                start_idx = (st.session_state[page_key] - 1) * items_per_page
                end_idx = min(start_idx + items_per_page, len(all_items))
                page_items = all_items[start_idx:end_idx]
                
                st.markdown("---")
                st.markdown(f"### 📋 {start_idx + 1} ~ {end_idx} 번째 항목")
                
                # 일괄 입력 폼
                with st.form(key=f"form_{mk}_{st.session_state[page_key]}"):
                    # 헤더
                    col_h1, col_h2, col_h3, col_h4, col_h5 = st.columns([2, 2, 1, 1.5, 1])
                    with col_h1:
                        st.markdown("**항목명**")
                    with col_h2:
                        st.markdown("**규격**")
                    with col_h3:
                        st.markdown("**수량**")
                    with col_h4:
                        st.markdown("**1일 작업량**")
                    with col_h5:
                        st.markdown("**단위**")
                    
                    st.markdown("---")
                    
                    # 입력 폼
                    form_inputs = {}
                    for i, item in enumerate(page_items):
                        manual_key = item["manual_key"]
                        
                        # 기존 값
                        existing_val = 0.0
                        existing_unit = item.get('unit', '') + "/일"
                        if manual_key in st.session_state["manual_rates"]:
                            existing_val = st.session_state["manual_rates"][manual_key].get("daily", 0)
                            existing_unit = st.session_state["manual_rates"][manual_key].get("unit", existing_unit)
                        
                        col1, col2, col3, col4, col5 = st.columns([2, 2, 1, 1.5, 1])
                        with col1:
                            st.text(item["name"])
                        with col2:
                            st.text(item.get("spec", ""))
                        with col3:
                            st.text(f"{item.get('qty', 0):,.1f}")
                        with col4:
                            daily = st.number_input(
                                "daily",
                                min_value=0.0,
                                value=float(existing_val),
                                step=0.1,
                                key=f"in_{mk}_{start_idx + i}",
                                label_visibility="collapsed"
                            )
                        with col5:
                            unit_in = st.text_input(
                                "unit",
                                value=existing_unit,
                                key=f"un_{mk}_{start_idx + i}",
                                label_visibility="collapsed"
                            )
                        
                        form_inputs[manual_key] = {"daily": daily, "unit": unit_in}
                    
                    st.markdown("---")
                    
                    # 일괄 저장 버튼
                    submitted = st.form_submit_button("💾 이 페이지 일괄 저장", width="stretch", type="primary")
                    
                    if submitted:
                        saved = 0
                        for mk_key, vals in form_inputs.items():
                            if vals["daily"] > 0:
                                st.session_state["manual_rates"][mk_key] = vals
                                saved += 1
                        st.success(f"✅ {saved}개 항목 저장 완료!")
                        st.rerun(scope="fragment")
            
            # 각 sub_tab에서 fragment 함수 호출
            for sub_tab, mk in zip(sub_tabs, sub_tab_keys):
                with sub_tab:
                    render_manual_input_page(mk, group_data[mk])