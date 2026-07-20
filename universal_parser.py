# -*- coding: utf-8 -*-
"""
universal_parser.py — 설정(config) 기반 범용 내역서/단가근거 파서

배경
----
회사·적산 소프트웨어마다 설계내역서 엑셀 양식이 다르다 (시트명, 열 배치,
계층 표시 방식, 단가근거 연결 방식 등). 하지만 시공량 산식(Q=/HR)은 전부
같은 국가 표준품셈 방법론을 쓰기 때문에 **완전히 동일**하다.

그래서 이 모듈은 "매번 다른 부분"만 TEMPLATES 설정으로 분리하고,
"항상 같은 부분"(Q값 추출 정규식)은 하나의 엔진으로 공유한다.
새 회사 양식을 만나면 파이썬 코드를 새로 짜지 않고, TEMPLATES에
설정 딕셔너리 하나를 추가하는 것으로 대응하는 게 목표다.

검증 상태
---------
아래 2개 템플릿은 실제 파일로 검증 완료. 결과가 기존 전용 파서
(hopyo_parser.py, naeyeok_parser.py)와 정확히 일치함을 확인함:
  - "표준형(산근호표)": 3_토목도급-동부지구... 파일
  - "코드매칭형(내역서산근)": (토목공사)화성 관로R18... 파일

새 템플릿 추가 방법
-------------------
1. 새 엑셀의 시트명/열 배치를 실측 확인 (openpyxl로 몇 행 찍어보기)
2. 아래 TEMPLATES 리스트에 dict 하나 추가:
   - item_sheet_names: 항목 시트 이름 후보 리스트
   - unit_price_sheet_names: 단가근거 시트 이름 후보 리스트
   - name_col, spec_col, qty_col, unit_col: 항목시트 열 인덱스
   - leaf_strategy: "column_tag"(특정 열에 값 있으면 리프) 또는
                    "regex_hierarchy"(gong_jong 열이 특정 패턴 아니면 리프)
   - hierarchy_strategy: "major_number_prefix"(들여쓰기0 "N.이름") 또는
                          "district_roman"(로마숫자+"N.N.N"코드)
   - link_strategy: "column_code"(특정 열 값이 코드) 또는
                    "text_regex"(정규식으로 텍스트에서 추출)
   - unit_price_code_col, unit_price_formula_col: 단가근거 시트 열 인덱스
3. python universal_parser.py <새엑셀경로> 로 검증 (자동판별 실패시 강제 지정 가능)
"""

from __future__ import annotations

import re
import sys
from typing import Dict, List, Optional, Tuple, Any


# ---------------------------------------------------------------------------
# 공용 Q값 추출 엔진 (모든 템플릿이 동일하게 재사용 — 국가 표준품셈 방법론이 같으므로)
# ---------------------------------------------------------------------------

def _build_q_regexes(unit_char_class: str):
    """템플릿별 단위 문자클래스로 정방향/역수 Q정규식을 생성."""
    fwd = re.compile(r"=\s*([\d,.]+)\s*([" + unit_char_class + r"]+)/hr", re.IGNORECASE)
    inv = re.compile(r"=\s*([\d,.]+)\s*hr/([" + unit_char_class + r"]+)", re.IGNORECASE)
    return fwd, inv


# 기본값(하위호환용) — 실제로는 항상 템플릿의 unit_char_class를 통해 생성해서 씀
_Q_FORWARD_RE, _Q_INVERSE_RE = _build_q_regexes(r"가-힣㎡㎥")
_HR_TO_DAY = 8


def _to_float(s: str) -> float:
    return float(s.replace(",", ""))


def _extract_q_from_block(rows, start: int, end: int, formula_col: int, q_forward_re=None, q_inverse_re=None):
    """[start, end) 범위에서 첫 Q=/HR(정방향) 또는 hr/단위(역수) 값을 찾아 (daily, unit) 반환."""
    q_forward_re = q_forward_re or _Q_FORWARD_RE
    q_inverse_re = q_inverse_re or _Q_INVERSE_RE
    found_fwd = None
    found_inv = None
    for j in range(start, end):
        row = rows[j]
        cell = row[formula_col] if len(row) > formula_col else None
        if not isinstance(cell, str):
            continue
        if found_fwd is None:
            m = q_forward_re.search(cell)
            if m:
                found_fwd = (_to_float(m.group(1)), m.group(2))
        if found_fwd is None and found_inv is None:
            mi = q_inverse_re.search(cell)
            if mi:
                found_inv = (_to_float(mi.group(1)), mi.group(2))
        if found_fwd is not None:
            break
    if found_fwd is not None:
        val, unit = found_fwd
        return round(val * _HR_TO_DAY, 4), unit
    if found_inv is not None:
        val, unit = found_inv
        if val > 0:
            return round(_HR_TO_DAY / val, 4), unit
    return None


# ---------------------------------------------------------------------------
# 템플릿 레지스트리
# ---------------------------------------------------------------------------

TEMPLATES: List[Dict[str, Any]] = [
    {
        "id": "standard_hopyo",
        "name": "표준형(산근호표)",
        "item_sheet_names": ["설계내역서"],
        "unit_price_sheet_names": ["단가산출근거"],
        "name_col": 1,
        "spec_col": 2,
        "qty_col": 3,
        "unit_col": 4,
        "leaf_strategy": "regex_hierarchy",
        "hierarchy_strategy": "district_roman",
        "gong_jong_col": 0,
        "hierarchy_leaf_regex": None,  # 로마숫자/3단계코드가 아니고 name이 있으면 리프 (district_roman 내부 처리)
        "link_strategy": "text_regex",
        "link_regex": r"산근\s*(\d+)\s*호표",
        "link_key_type": "int",
        "unit_price_block_start_col": 1,
        "unit_price_block_start_regex": r"제\s*(\d+)\s*호표",
        "unit_price_formula_col": 1,
        "unit_char_class": r"가-힣㎡㎥",
    },
    {
        "id": "code_match_naeyeok",
        "name": "코드매칭형(내역서산근)",
        "item_sheet_names": ["내역서"],
        "unit_price_sheet_names": ["일위대가_산근"],
        "name_col": 0,
        "spec_col": 1,
        "qty_col": 2,
        "unit_col": 3,
        "leaf_strategy": "column_tag",
        "leaf_tag_col": 12,
        "hierarchy_strategy": "major_number_prefix",
        "line_marker_symbol": "■",
        "line_spec_col": 1,
        "link_strategy": "column_code",
        "link_code_col": 17,
        "link_key_type": "str",
        "unit_price_block_start_col": 7,
        "unit_price_block_start_regex": None,  # 코드매칭형은 정규식 없이 값 존재 자체가 블록시작
        "unit_price_formula_col": 0,
        "unit_char_class": r"가-힣㎡㎥A-Za-z0-9",
    },
]


def detect_template(wb) -> Optional[Dict[str, Any]]:
    """워크북 시트명으로 등록된 템플릿 중 매칭되는 것을 찾는다."""
    if wb is None:
        return None
    names = set(wb.sheetnames)
    for tmpl in TEMPLATES:
        if any(s in names for s in tmpl["item_sheet_names"]) and \
           any(s in names for s in tmpl["unit_price_sheet_names"]):
            return tmpl
    return None


def _get_sheet(wb, name_candidates):
    for n in name_candidates:
        if n in wb.sheetnames:
            return wb[n]
    return None


# ---------------------------------------------------------------------------
# 단가근거 시트 파싱 (공용 — 블록 탐지 방식만 템플릿마다 다름)
# ---------------------------------------------------------------------------

def parse_unit_price_generic(wb, tmpl: Dict[str, Any]) -> Dict[Any, Tuple[float, str]]:
    """
    단가근거 시트에서 {키(호표번호 또는 코드): (일작업량, 단위)} 추출.
    블록 시작 판별은 템플릿의 unit_price_block_start_col/regex를 따르고,
    Q값 추출 자체는 모든 템플릿이 공유하는 _extract_q_from_block을 쓴다.
    """
    ws = _get_sheet(wb, tmpl["unit_price_sheet_names"])
    if ws is None:
        return {}

    rows = [r for r in ws.iter_rows(values_only=True)]
    block_col = tmpl["unit_price_block_start_col"]
    block_regex = tmpl.get("unit_price_block_start_regex")
    formula_col = tmpl["unit_price_formula_col"]
    key_type = tmpl.get("link_key_type", "str")
    unit_char_class = tmpl.get("unit_char_class", r"가-힣㎡㎥")
    q_fwd_re, q_inv_re = _build_q_regexes(unit_char_class)

    # 1) 블록 시작행 인덱싱
    starts: List[Tuple[Any, int]] = []
    seen_keys = set()
    if block_regex:
        pat = re.compile(block_regex)
        for i, r in enumerate(rows):
            c = r[block_col] if len(r) > block_col else None
            if isinstance(c, str):
                m = pat.search(c)
                if m:
                    key = int(m.group(1)) if key_type == "int" else m.group(1)
                    if key not in seen_keys:
                        seen_keys.add(key)
                        starts.append((key, i))
    else:
        for i, r in enumerate(rows):
            c = r[block_col] if len(r) > block_col else None
            if isinstance(c, str) and c.strip():
                key = c.strip()
                if key not in seen_keys:
                    seen_keys.add(key)
                    starts.append((key, i))

    if not starts:
        return {}

    # 2) 블록별 Q값 추출
    result: Dict[Any, Tuple[float, str]] = {}
    for idx, (key, s) in enumerate(starts):
        end = starts[idx + 1][1] if idx + 1 < len(starts) else len(rows)
        found = _extract_q_from_block(rows, s, end, formula_col, q_fwd_re, q_inv_re)
        if found:
            result[key] = found
    return result


# ---------------------------------------------------------------------------
# 항목 시트 파싱 (계층 전략별 분기)
# ---------------------------------------------------------------------------

def _parse_items_major_number_prefix(ws, tmpl: Dict[str, Any]) -> List[Dict]:
    """'N. 이름'(들여쓰기0) 대공종 + '■'라인 추적. leaf_strategy='column_tag' 전제."""
    major_re = re.compile(r"^\d+\.\s*(.+)$")
    name_col = tmpl["name_col"]
    spec_col = tmpl["spec_col"]
    qty_col = tmpl["qty_col"]
    unit_col = tmpl["unit_col"]
    leaf_tag_col = tmpl["leaf_tag_col"]
    line_symbol = tmpl["line_marker_symbol"]
    line_spec_col = tmpl["line_spec_col"]
    link_code_col = tmpl.get("link_code_col")

    current_major = None
    current_line = ""
    items: List[Dict] = []

    for row in ws.iter_rows(values_only=True):
        raw_name = row[name_col] if len(row) > name_col else None
        if not raw_name:
            continue
        s = str(raw_name)
        stripped = s.lstrip(" ")
        indent = len(s) - len(stripped)

        tag = row[leaf_tag_col] if len(row) > leaf_tag_col else None
        qty = row[qty_col] if len(row) > qty_col else None
        is_leaf = tag not in (None, "") and isinstance(qty, (int, float))

        if not is_leaf and indent == 0:
            if stripped.startswith(line_symbol):
                spec = str(row[line_spec_col]).strip() if len(row) > line_spec_col and row[line_spec_col] else ""
                current_line = spec
            else:
                m = major_re.match(stripped)
                if m:
                    current_major = m.group(1).strip()
            continue

        if is_leaf:
            spec = str(row[spec_col]).strip() if len(row) > spec_col and row[spec_col] else ""
            unit = str(row[unit_col]).strip() if len(row) > unit_col and row[unit_col] else ""
            code = row[link_code_col] if link_code_col is not None and len(row) > link_code_col else None
            code = code.strip() if isinstance(code, str) else None
            items.append({
                "name": stripped, "spec": spec, "qty": qty, "unit": unit,
                "code": code, "category": current_major, "line": current_line,
            })
    return items


def _parse_items_district_roman(ws, tmpl: Dict[str, Any]) -> List[Dict]:
    """
    로마숫자 지구 + 'N.N.N' 3단계 코드(대분류) + 'N)' 소분류 + '(N) #N...' 세부구분자.

    app.py 레거시 계층파서를 그대로 이식한 상태머신. 단순 flat 추출이 아니라
    아래 병합 규칙까지 정확히 재현해야 한다 (실측으로 차이 확인됨):
      - sub_category(예: '1) 토공')는 같은 level+name+district 조합이면
        재사용(reuse)된다 — 시트 내 여러 곳에 흩어져 나와도 하나로 누적.
      - 항목은 현재 활성 컨테이너(sub_sub > sub > category) 안에서
        (name, spec)이 같으면 수량을 합산한다.
      - 예외: '#N' 형태 세부구분자(sub_sub_category)이고 그 부모 sub_category
        이름에 '추진'이 포함되면, sub_sub 레벨을 건너뛰고 부모 sub_category
        레벨로 합산한다 (여러 #N 추진 구간의 같은 항목을 하나로 합치기 위함).
    내부적으로 중첩 hierarchy를 만들어 정확히 합산한 뒤, 최종적으로
    {name,spec,qty,unit,code,category,line} 평평한 리스트로 변환해 반환한다.
    """
    roman_nums = ['Ⅰ', 'Ⅱ', 'Ⅲ', 'Ⅳ', 'Ⅴ', 'Ⅵ', 'Ⅶ', 'Ⅷ', 'Ⅸ', 'Ⅹ']
    major_code_re = re.compile(r"^\d+\.\d+\.\d+$")
    sub_re = re.compile(r"^\d+\)$")
    hash_paren_re = re.compile(r"^\(\d+\)$")
    hash_name_re = re.compile(r"^#\d+")
    hash_gj_re = re.compile(r"^#\d+")

    gong_jong_col = tmpl["gong_jong_col"]
    name_col = tmpl["name_col"]
    spec_col = tmpl["spec_col"]
    qty_col = tmpl["qty_col"]
    unit_col = tmpl["unit_col"]
    link_regex = re.compile(tmpl["link_regex"]) if tmpl.get("link_regex") else None
    key_type = tmpl.get("link_key_type", "str")

    hierarchy: List[Dict] = []
    current_district = None
    current_category = None
    current_sub_category = None
    current_sub_sub_category = None

    def _merge_or_append(container_items: List[Dict], item: Dict):
        existing = next((i for i in container_items
                          if i['name'] == item['name'] and i.get('spec') == item.get('spec')), None)
        if existing:
            existing['qty'] = existing.get('qty', 0) + item.get('qty', 0)
        else:
            container_items.append(item)

    for row in ws.iter_rows(values_only=True):
        gj = row[gong_jong_col] if len(row) > gong_jong_col else None
        gj = str(gj).strip() if gj else ""
        name = row[name_col] if len(row) > name_col else None
        name = str(name).strip() if name else ""

        if gj in roman_nums:
            current_district = gj
            current_sub_category = None
            current_sub_sub_category = None
            continue

        if major_code_re.match(gj):
            if current_category:
                if current_sub_category:
                    current_category['sub_categories'].append(current_sub_category)
                    current_sub_category = None
                if current_category.get('items') or current_category.get('sub_categories'):
                    hierarchy.append(current_category)
            current_category = {'level': gj, 'name': name, 'items': [], 'sub_categories': []}
            current_sub_category = None
            continue

        is_hash_separator = bool(
            (hash_paren_re.match(gj) and name and hash_name_re.match(name)) or hash_gj_re.match(gj)
        )
        if is_hash_separator:
            if current_sub_category:
                current_sub_sub_category = {'level': gj, 'name': name, 'district': current_district, 'items': []}
                current_sub_category.setdefault('sub_categories', []).append(current_sub_sub_category)
            continue

        if sub_re.match(gj):
            if current_category:
                if current_sub_sub_category and current_sub_category:
                    if not any(s is current_sub_sub_category for s in current_sub_category['sub_categories']):
                        current_sub_category.setdefault('sub_categories', []).append(current_sub_sub_category)
                current_sub_sub_category = None

                existing_sub = next((s for s in current_category['sub_categories']
                                      if s['level'] == gj and s['name'] == name and s.get('district') == current_district), None)
                if existing_sub:
                    if current_sub_category and current_sub_category is not existing_sub:
                        if not any(s is current_sub_category for s in current_category['sub_categories']):
                            current_category['sub_categories'].append(current_sub_category)
                    current_sub_category = existing_sub
                else:
                    if current_sub_category:
                        if not any(s is current_sub_category for s in current_category['sub_categories']):
                            current_category['sub_categories'].append(current_sub_category)
                    current_sub_category = {'level': gj, 'name': name, 'items': [], 'sub_categories': [], 'district': current_district}
            continue

        # 항목 행: gong_jong 비어있고 name 있음
        if current_category and not gj and name:
            qty_val = row[qty_col] if len(row) > qty_col else None
            unit_val = row[unit_col] if len(row) > unit_col else None
            try:
                qty = float(qty_val) if qty_val else 0
            except (TypeError, ValueError):
                qty = 0
            if qty <= 0:
                continue
            spec = str(row[spec_col]).strip() if len(row) > spec_col and row[spec_col] else ""
            unit = str(unit_val).strip() if unit_val else ""

            code = None
            if link_regex:
                for v in row:
                    if isinstance(v, str):
                        m = link_regex.search(v)
                        if m:
                            code = int(m.group(1)) if key_type == "int" else m.group(1)
                            break

            item = {'name': name, 'spec': spec, 'qty': qty, 'unit': unit, 'district': current_district, 'code': code}

            if (current_sub_sub_category and current_sub_category and
                    "추진" in current_sub_category.get('name', '') and
                    hash_name_re.match(current_sub_sub_category.get('name', ''))):
                _merge_or_append(current_sub_category['items'], item)
            elif current_sub_sub_category:
                _merge_or_append(current_sub_sub_category['items'], item)
            elif current_sub_category:
                _merge_or_append(current_sub_category['items'], item)
            else:
                _merge_or_append(current_category['items'], item)

    if current_category:
        if current_sub_sub_category and current_sub_category:
            if not any(s is current_sub_sub_category for s in current_sub_category['sub_categories']):
                current_sub_category.setdefault('sub_categories', []).append(current_sub_sub_category)
        if current_sub_category:
            if not any(s is current_sub_category for s in current_category['sub_categories']):
                current_category['sub_categories'].append(current_sub_category)
        if current_category.get('items') or current_category.get('sub_categories'):
            hierarchy.append(current_category)

    # 중첩 hierarchy → 평평한 items 리스트로 변환 (category=최상위 대분류 이름, line=지구)
    def _strip_major_prefix(name: str) -> str:
        m = re.match(r"^\d+\.\s*(.+)$", name)
        return m.group(1).strip() if m else name

    flat_items: List[Dict] = []

    def _walk(container, top_category_name, district):
        for it in container.get('items', []):
            flat_items.append({
                'name': it['name'], 'spec': it.get('spec', ''), 'qty': it.get('qty', 0),
                'unit': it.get('unit', ''), 'code': it.get('code'),
                'category': top_category_name, 'line': it.get('district', district),
            })
        for sub in container.get('sub_categories', []):
            _walk(sub, top_category_name, sub.get('district', district))

    for cat in hierarchy:
        top_name = _strip_major_prefix(cat['name'])
        _walk(cat, top_name, None)

    return flat_items


def parse_items_generic(wb, tmpl: Dict[str, Any]) -> List[Dict]:
    """템플릿의 hierarchy_strategy에 따라 적절한 파서로 위임."""
    ws = _get_sheet(wb, tmpl["item_sheet_names"])
    if ws is None:
        return []
    strategy = tmpl["hierarchy_strategy"]
    if strategy == "major_number_prefix":
        return _parse_items_major_number_prefix(ws, tmpl)
    elif strategy == "district_roman":
        return _parse_items_district_roman(ws, tmpl)
    else:
        raise ValueError(f"알 수 없는 hierarchy_strategy: {strategy}")


def parse_with_template(wb, tmpl: Dict[str, Any]):
    """편의 함수: 항목 + 단가근거를 한번에 파싱."""
    items = parse_items_generic(wb, tmpl)
    unit_prices = parse_unit_price_generic(wb, tmpl)
    return items, unit_prices


def parse_auto(wb):
    """템플릿 자동판별 후 파싱. 매칭 실패시 (None, [], {}) 반환."""
    tmpl = detect_template(wb)
    if tmpl is None:
        return None, [], {}
    items, unit_prices = parse_with_template(wb, tmpl)
    return tmpl, items, unit_prices


# ---------------------------------------------------------------------------
# CLI 검증
# ---------------------------------------------------------------------------

def _main(argv):
    import warnings
    warnings.filterwarnings("ignore")
    from openpyxl import load_workbook

    if len(argv) < 2:
        print("사용법: python universal_parser.py <엑셀경로> [템플릿id 강제지정]", file=sys.stderr)
        return 2
    path = argv[1]
    force_id = argv[2] if len(argv) > 2 else None

    print(f"[로드] {path}")
    wb = load_workbook(path, read_only=True, data_only=True)

    if force_id:
        tmpl = next((t for t in TEMPLATES if t["id"] == force_id), None)
        if tmpl is None:
            print(f"알 수 없는 템플릿id: {force_id}")
            return 2
    else:
        tmpl = detect_template(wb)

    if tmpl is None:
        print("❌ 매칭되는 템플릿이 없습니다. (신규 양식 — TEMPLATES에 추가 필요)")
        print(f"   시트 목록: {wb.sheetnames}")
        return 1

    print(f"[템플릿] {tmpl['name']} ({tmpl['id']})")
    items, unit_prices = parse_with_template(wb, tmpl)
    print(f"[파싱] 항목 {len(items)}개, 단가근거 {len(unit_prices)}개")

    matched = [it for it in items if it.get("code") in unit_prices]
    rate = (len(matched) / len(items) * 100) if items else 0.0
    print(f"[매칭] {len(matched)}개 ({rate:.1f}%)")

    from collections import Counter
    cat_counts = Counter(it.get("category") for it in items if it.get("category"))
    print("[대공종별 항목수]")
    for cat, cnt in cat_counts.most_common(15):
        print(f"    {cat:15} {cnt}개")

    return 0


if __name__ == "__main__":
    sys.exit(_main(sys.argv))